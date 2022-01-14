import datetime
import os
import random
import time
import urllib

import telebot
from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler
import urllib.request

import Data
import Notifications
import Other_function
import SQLite
import What_i_can_do

tconv = lambda x: time.strftime("%d.%m.%Y %H:%M:%S", time.localtime(x))  # Конвертация даты в читабельный вид
bot = Data.bot
answer_bot = 'Бот ответил:\n'


# Получаем имя пользователя: Администратор/Пользователь + Имя + ID
def full_name_user(message):
    check_admin = SQLite.check_for_admin(message.from_user.id)  # Проверяем является ли пользователь админом
    if check_admin is True:
        status_user = 'Администратор '
    else:
        status_user = 'Пользователь '
    name_user = message.from_user.first_name + ' (ID: ' + str(message.from_user.id) + ') '  # Получаем имя и id
    pattern = str(tconv(message.date)) + '\n' + status_user + name_user  # Итог дата, /n, статус и данные пользователя
    return pattern


# Приветственное сообщение
@bot.message_handler(commands=['start'])
def start_command(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is False:  # Если пользователь отсутствует в БД
        # Приветственное сообщение
        hello_message = 'Добро пожаловать ' + message.from_user.first_name + '!\n' + \
                        'Это информационный бот IT отдела. Тут можно узнать кто из системных администраторов ' \
                        'дежурный в ближайшие дни, кто и когда отсутствует и прочая информация.\n' + \
                        'Для того чтобы пользоваться функциями бота, необходимо пройти регистрацию нажав /register. ' \
                        'Тем самым вы даёте согласие на хранение и обработку данных о вашем аккаунте. В базу данных ' \
                        'будут занесены следующие сведения:\n ' + \
                        'ID: ' + str(message.from_user.id) + '\n' + \
                        'Имя: ' + str(message.from_user.first_name) + '\n' + \
                        'Фамилия: ' + str(message.from_user.last_name) + '\n' + \
                        'Username:  @' + str(message.from_user.username) + '\n'
        Data.bot.send_message(message.from_user.id, hello_message)
        print(answer_bot + hello_message + '\n')
    else:
        end_text = 'Привет еще раз, ' + message.from_user.first_name + '\n' + 'Мы уже знакомы!'  # Эта строка
        # появится если уже зарегистрированный пользователь попытается заново пройти регистрацию
        Data.bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


#  Регистрация данных о пользователе в БД
@bot.message_handler(commands=['register'])
def register(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is False:  # Если пользователь отсутствует в БД
        SQLite.db_table_val(message)
        time.sleep(5)  # Подождать указанное кол-во секунд
        register_message = 'Добро пожаловать ' + message.from_user.first_name + '!\n' + \
                           'Регистрация успешно завершена!' + '\n' + \
                           'Чтобы узнать, что умеет бот, жми /help.\n' \
                           'Не забудь подписаться на рассылку, чтобы быть в курсе последних событий, жми /subscribe'
        Data.bot.send_message(message.from_user.id, register_message)  # Бот пришлёт уведомление об успешной регистрации
        print(answer_bot + register_message + '\n')
    else:  # Иначе бот уведомит о том что пользователь уже регистрировался
        end_text = 'Вы уже зарегистрированы!' + '\n' + \
                   'Чтобы узнать что умеет бот жми /help.'
        Data.bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


#  Удаление данных о пользователе из БД
@bot.message_handler(commands=['log_out'])
def log_out(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Если пользователь присутствует в БД
        SQLite.log_out(message.from_user.id)  # Удаление данных из БД
        time.sleep(5)  # Подождать указанное кол-во секунд
        log_out_message = 'До новых встреч ' + message.from_user.first_name + '!\n' + \
                          'Данные о вашем аккаунте успешно удалены!' + '\n' + \
                          'Чтобы снова пользоваться функционалом бота, жми /register.'
        Data.bot.send_message(message.from_user.id, log_out_message)  # Прощальное сообщение
        print(answer_bot + log_out_message + '\n')
    else:  # Иначе бот уведомит о том что пользователь ещё не регистрировался
        end_text = 'Нельзя удалить данные которых нет :)\n' + 'Чтобы это сделать, нужно зарегистрироваться!'
        Data.bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


#  Список доступных команд
@bot.message_handler(commands=['help'])
def help_command(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        keyboard = telebot.types.InlineKeyboardMarkup()  # Вызов кнопки
        keyboard.add(telebot.types.InlineKeyboardButton('Написать разработчику', url='t.me/nikita_it_remit'))
        bot.send_message(message.chat.id, What_i_can_do.can_help(message), reply_markup=keyboard)  # Показ списка
        # доступных команд и кнопки "Написать разработчику"
        print(answer_bot + What_i_can_do.can_help(message) + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['invent'])
def invent(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    start = time.time()  # Засекает время начала выполнения скрипта
    list_name = 'Инвентаризация'  # Получаем имя страницы по ключу

    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Проверка админ ли юзер
            date_list_today = Other_function.read_sheet(list_name)[1]
            event_data = date_list_today[0]
            first_date = event_data[0]
            first_date_format = first_date.strftime("%d.%m.%Y")
            event = event_data[1]
            name_from_SQL = SQLite.get_user_info(Other_function.get_key(Data.user_data, event))
            if name_from_SQL is None:
                name_from_SQL = event
            date_now = datetime.datetime.now()  # Получаем текущую дату
            difference_date = first_date - date_now
            difference_date = difference_date.days + 1
            print(name_from_SQL)

            # Склоняем "день"
            def count_day():
                if difference_date == 0:
                    return 'Сегодня инвентаризация.'
                elif difference_date == 1:
                    return 'До предстоящей инвентаризации остался 1 день.'
                elif 1 < difference_date <= 4:
                    return 'До предстоящей инвентаризации осталось ' + str(difference_date) + ' дня.'
                elif difference_date == 5:
                    return 'До предстоящей инвентаризации осталось 5 дней.'
                elif difference_date > 5:
                    return 'Следующая инвентаризация состоится ' + str(first_date_format) + '.'

            text_day = count_day()  # Кол-во дней до инвентаризации
            text_who = 'Судя по графику, выходит ' + name_from_SQL + '.'  # Имя следующего дежурного
            end_text = text_day + '\n' + text_who  # Объединяем строки выше в одну
            # Если в БД у пользователя содержится стикер
            if SQLite.get_user_sticker(Other_function.get_key(Data.user_data, event)) is not None:
                # Пришлёт сообщение о дежурном
                bot.send_message(message.chat.id, end_text)
                # Пришлёт стикер этого дежурного
                bot.send_sticker(message.chat.id,
                                 SQLite.get_user_sticker(Other_function.get_key(Data.user_data, event)))
            else:
                # Пришлёт сообщение о дежурном
                bot.send_message(message.chat.id, end_text)
        else:  # Если юзер не админ, он получит следующее сообщение
            end_text = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.chat.id, end_text)
    else:  # Если пользователь не зарегистрирован, он получит следующее сообщение
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)

    end = time.time()  # Засекает время окончания скрипта
    print(answer_bot + end_text + '\n' + 'Время работы запроса(сек): ' + str(int(end - start)) + '\n')


#  Получить случайное имя из сисадминов
@bot.message_handler(commands=['random'])
def random_name(message):
    user_id = message.from_user.id
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(user_id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(user_id) is True:  # Проверка админ ли юзер
            list_name = ['Паша', 'Дима', 'Никита']  # Список имён
            r_name = random.choice(list_name)  # Получение случайного значения из списка
            bot.send_message(user_id, text=r_name)  # Отправка сообщения с рандомным именем
            print(answer_bot + r_name + '\n')
        else:  # Если пользователь не админ бот уведомит об этом
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(user_id, text_message)
            print(answer_bot + text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(user_id, end_text)
        print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['set_admin'])
def set_to_admin(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Если пользователь админ
            text_message = 'Чтобы назначить администратора, перешли мне сообщение от этого человека\n'
            bot.send_message(message.from_user.id, text=text_message)  # Бот пришлёт выше указанный текст
            print(text_message + '\n')
            bot.register_next_step_handler(message, receive_id)  # Регистрация следующего действия
        else:  # Если пользователь не админ, бот сообщит об этом
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.from_user.id, text_message)
            print(text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def receive_id(message):
    try:
        id_future_admin = message.forward_from.id  # Получаем id человека полученного из пересылаемого сообщения
        first_name_future_admin = str(message.forward_from.first_name)  # Получаем имя будущего админа
        last_name_future_admin = str(message.forward_from.last_name)  # Получаем фамилию будущего админа
        full_name_future_admin = first_name_future_admin + ' ' + last_name_future_admin  # Склеиваем данные воедино
        print(full_name_user(message) + ' переслал сообщение от пользователя ' + full_name_future_admin +
              ' содержащее текст:\n' + message.text)
        answer_text = 'Пользователь <' + full_name_future_admin + '> добавлен в список администраторов'
        # msg = bot.send_message(chat_id, answer_text)
        if SQLite.check_for_existence(id_future_admin) is True:  # Проверка на наличие человека в БД
            if SQLite.check_for_admin(id_future_admin) is False:  # Проверка админ ли юзер
                SQLite.update_sqlite_table('admin', id_future_admin, 'status')  # Обновляем статус нового админа в БД
                bot.send_message(message.from_user.id, answer_text)  # Бот уведомляет об этом того кто выполнил запрос
                print(answer_text + '\n')
                bot.send_message(id_future_admin, 'Администратор <' + message.from_user.first_name +
                                 '> предоставил вам права администратора')  # Бот уведомляет пользователя, что
                # такой-то юзер, дал ему права админа
            else:  # Если тот кому предоставляют права уже админ, бот сообщит об ошибке
                end_text = 'Нельзя пользователю присвоить статус <admin> поскольку он им уже является'
                bot.send_message(message.from_user.id, end_text)
        else:  # Если того кому пытаются дать права нет в БД, бот сообщит об ошибке
            end_text = 'Вы пытаетесь дать права администратора пользователю который отсутствует в базе данных!'
            bot.send_message(message.chat_id, end_text)
            print(end_text + '\n')
    except Exception as error:  # В любом другом случае бот сообщит об ошибке
        bot.reply_to(message, 'Что-то пошло не так. Чтобы попробовать снова, жми /set_admin')
        print(str(error))
        Other_function.logging_event('error', str(error))

    return


@bot.message_handler(commands=['set_user'])
def set_to_user(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Если пользователь админ
            text_message = 'Чтобы пользователю присвоить статус <user>, перешли мне сообщение от этого человека\n'
            bot.send_message(message.from_user.id, text=text_message)  # Бот пришлёт выше указанный текст
            print(answer_bot + text_message + '\n')
            bot.register_next_step_handler(message, receive_id_user)  # Регистрация следующего действия
        else:  # Если пользователь не админ, бот сообщит об этом
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.from_user.id, text_message)
            print(text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def receive_id_user(message):
    try:
        chat_id = message.chat.id  # Получаем id чата
        id_future_user = message.forward_from.id  # Получаем id человека полученного из пересылаемого сообщения
        first_name_future_user = str(message.forward_from.first_name)  # Получаем имя будущего юзера
        last_name_future_user = str(message.forward_from.last_name)  # Получаем фамилию будущего юзера
        full_name_future_user = first_name_future_user + ' ' + last_name_future_user  # Склеиваем данные воедино
        print(full_name_user(message) + ' переслал сообщение от пользователя ' + full_name_future_user +
              ' содержащее текст:\n' + message.text)
        answer_text = 'Пользователю <' + full_name_future_user + '> присвоен статус <user>'
        if SQLite.check_for_existence(id_future_user) is True:  # Проверка на наличие человека в БД
            if SQLite.check_for_admin(id_future_user) is True:  # Проверка админ ли юзер
                SQLite.update_sqlite_table('user', id_future_user, 'status')  # Обновляем статус нового юзера в БД
                bot.send_message(message.from_user.id, answer_text)  # Бот уведомляет об этом того кто выполнил запрос
                print(answer_text + '\n')
                bot.send_message(id_future_user, 'Администратор <' + message.from_user.first_name +
                                 '> лишил вас прав администратора')  # Бот уведомляет нового юзера, что
                # пользователь <Имя>, лишил его прав админа
            else:  # Если тот, кого лишают прав админа, уже и так юзер, бот сообщит об ошибке
                end_text = 'Нельзя пользователю присвоить статус <user> поскольку он им уже является'
                bot.send_message(message.from_user.id, end_text)
        else:  # Если того, кого пытаются лишить прав админа, нет в БД, бот сообщит об ошибке
            end_text = 'Вы пытаетесь присвоить пользователю статус <user>, который отсутствует в базе данных!'
            bot.send_message(chat_id, end_text)
            print(end_text + '\n')

    except Exception as error:  # В любом другом случае бот сообщит об ошибке
        bot.reply_to(message, 'Что-то пошло не так. Чтобы попробовать снова, жми /set_user')
        print(str(error))
        Other_function.logging_event('error', str(error))
    return


@bot.message_handler(commands=['subscribe'])
def set_subscribe(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_notification(message.from_user.id) is False:  # Если пользователь не подписчик
            SQLite.update_sqlite_table('yes', message.from_user.id, 'notification')  # Присвоить статус <подписан>
            end_text = 'Вы подписаны на уведомления. Теперь вам будут приходить уведомления о том кто дежурит в ' \
                       'выходные, кто в отпуске и прочая информация.\n Чтобы отписаться жми /unsubscribe '
            bot.send_message(message.from_user.id, end_text)  # Отправка текста выше
            #  Отсылка уведомлений о действии разработчику
            bot.send_message(chat_id=Data.list_admins.get('Никита'), text=full_name_user(message) + 'подписался на '
                                                                                                    'уведомления.')
            print(answer_bot + end_text + '\n')
        else:  # Если подписчик пытается подписаться, то получит ошибку
            end_text = 'Вы уже подписаны на уведомления.'
            bot.send_message(message.from_user.id, end_text)
            print(answer_bot + end_text + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы управлять подпиской нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['unsubscribe'])
def set_subscribe(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_notification(message.from_user.id) is True:  # Если пользователь подписчик
            SQLite.update_sqlite_table('no', message.from_user.id, 'notification')  # Присвоить в БД статус <не
            # подписан>
            end_text = 'Рассылка отключена.\n Чтобы подписаться жми /subscribe'
            bot.send_message(message.from_user.id, end_text)  # Отправка текста выше
            #  Отсылка уведомлений о действии разработчику
            bot.send_message(chat_id=Data.list_admins.get('Никита'), text=full_name_user(message) + 'отписался от '
                                                                                                    'уведомлений.')
            print(answer_bot + end_text + '\n')
        else:  # Если не подписчик пытается отписаться, то получит ошибку
            end_text = 'Нельзя отказаться от уведомлений на которые не подписан.'
            bot.send_message(message.from_user.id, end_text)
            print(answer_bot + end_text + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы управлять подпиской нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['change_sticker'])
def change_sticker_1(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Если пользователь админ
            msg = bot.send_message(message.from_user.id, 'Отправь мне стикер который хочешь привязать в своей учётной '
                                                         'записи!')
            bot.register_next_step_handler(msg, change_sticker_2)  # Регистрация следующего действия
        else:  # Если пользователь не админ, бот сообщит об этом
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.from_user.id, text_message)
            print(text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def change_sticker_2(message):
    print(full_name_user(message) + 'отправил стикер ' + message.sticker.file_id)
    SQLite.update_sqlite_table(message.sticker.file_id, message.from_user.id, 'sticker')
    end_text = 'Стикер обновлён'
    bot.send_message(message.from_user.id, end_text)
    print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['dezhurnyj'])
def dej(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        text_message = 'Что вы хотите получить?'
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        buttons = ['Имя следующего дежурного', 'Список дежурных']
        keyboard.add(*buttons)
        bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)  #
        bot.register_next_step_handler(message, dej_step_2)  # Регистрация следующего действия
        print(answer_bot + text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def dej_step_2(message):
    print(full_name_user(message) + 'написал:\n' + message.text)
    date_list = Other_function.read_sheet('Дежурный')[0]
    count_date_list = len(date_list)
    hide_keyboard = telebot.types.ReplyKeyboardRemove()
    try:
        if message.text == 'Имя следующего дежурного':
            date_list_today = Other_function.read_sheet('Дежурный')[1]
            event_data = date_list_today[0]
            first_date = event_data[0]
            first_date = first_date.strftime("%d.%m.%Y")
            last_date = event_data[1]
            last_date = last_date.strftime("%d.%m.%Y")
            event = event_data[2]
            name_from_SQL = SQLite.get_user_info(Other_function.get_key(Data.user_data, event))
            text_message = 'В период с ' + first_date + ' по ' + last_date + ' ' + 'будет дежурить ' + name_from_SQL + \
                           '.'
            # Если в БД у пользователя содержится стикер
            if SQLite.get_user_sticker(Other_function.get_key(Data.user_data, event)) is not None:
                # Пришлёт сообщение о дежурном
                bot.send_message(message.chat.id, text_message, reply_markup=hide_keyboard)
                # Пришлёт стикер этого дежурного
                bot.send_sticker(message.chat.id,
                                 SQLite.get_user_sticker(Other_function.get_key(Data.user_data, event)))
            else:
                # Пришлёт сообщение о дежурном
                bot.send_message(message.chat.id, text_message, reply_markup=hide_keyboard)
            print(answer_bot + text_message + '\n')
        elif message.text == 'Список дежурных':
            def count_records(n):
                records = ['запись', 'записи', 'записей']

                if n == 0:
                    return 'К сожалению, нет данных о предстоящих дежурствах.'
                elif n % 10 == 1 and n % 100 != 11:  # <1, 21 запись>
                    r = 0
                    return 'На данный момент есть ' + str(n) + ' ' + records[r] + '. Сколько событий показать?'
                elif 2 <= n % 10 <= 4 and (n % 100 < 10 or n % 100 >= 20):  # <3, 22 записи>
                    r = 1
                    return 'На данный момент есть ' + str(n) + ' ' + records[r] + '. Сколько событий показать?'
                else:  # <5, 47 записей>
                    r = 2
                    return 'На данный момент есть ' + str(n) + ' ' + records[r] + '. Сколько событий показать?'

            text_message = count_records(count_date_list)
            # text_message = 'На данный момент доступно ' + str(count_date_list) + ' записей\n' + \
            #                'Сколько событий показать?'
            bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
            bot.register_next_step_handler(message, dej_step_3)  # Регистрация следующего действия
            print(answer_bot + text_message + '\n')
    except Exception as error:  # В любом другом случае бот сообщит об ошибке
        text_message = 'Что-то пошло не так. Чтобы попробовать снова, жми /test'
        bot.reply_to(message, text_message)
        print(answer_bot + text_message + '\n')
        print(str(error))
        Other_function.logging_event('error', str(error))


def dej_step_3(message):
    print(full_name_user(message) + 'написал:\n' + message.text)
    hide_keyboard = telebot.types.ReplyKeyboardRemove()
    try:
        count = int(message.text)
        date_list = Other_function.read_sheet('Дежурный')[0]

        if date_list is not None:
            if count <= len(date_list):
                print(answer_bot)
                for i in range(0, count):
                    event_data = date_list[i]
                    first_date = event_data[0]
                    first_date_format = first_date.strftime("%d.%m.%Y")
                    last_date = event_data[1]
                    last_date_format = last_date.strftime("%d.%m.%Y")
                    event = event_data[2]
                    text_message = 'В период с ' + first_date_format + ' по ' + last_date_format + ' будет дежурить ' \
                                   + event
                    bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
                    print(text_message)
            else:
                text_message = 'Вы запрашиваете ' + str(count) + ' записей, а есть только ' + str(
                    len(date_list)) + '.\n' + 'Попробуйте снова - /dezhurnyj'
                bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
                print(answer_bot + text_message + '\n')
    except Exception as error:  # В любом другом случае бот сообщит об ошибке
        text_message = 'Я не распознал число, попробуйте снова - /dezhurnyj'
        bot.reply_to(message, text_message)
        print(answer_bot + text_message + '\n')
        print(str(error))
        Other_function.logging_event('error', str(error))


@bot.message_handler(commands=['get_list'])
def get_list(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Если пользователь админ
            bot.send_message(message.from_user.id, SQLite.get_list_users())
        else:  # Если пользователь не админ, бот сообщит об этом
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.from_user.id, text_message)
            print(text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


@bot.message_handler(commands=['feed_back'])
def feed_back(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        text_message = 'Выберите тип обращения'
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        buttons = ['Что-то не работает', 'Есть идея новой функции', 'Другое']
        keyboard.add(*buttons)
        bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)  #
        bot.register_next_step_handler(message, feed_back_step_2)  # Регистрация следующего действия
        print(answer_bot + text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def feed_back_step_2(message):
    answer = 'Опишите суть обращения. Чем подробнее тем лучше.\n'
    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    buttons = ['Отмена']
    keyboard.add(*buttons)
    bot.send_message(message.from_user.id, answer, reply_markup=keyboard)  #
    contacting_technical_support = message.text + '\n'
    bot.register_next_step_handler(message, feed_back_step_3,
                                   contacting_technical_support)  # Регистрация следующего действия
    print(answer_bot + answer + '\n')


def feed_back_step_3(message, text_problem):
    hide_keyboard = telebot.types.ReplyKeyboardRemove()
    if message.text == 'Отмена':
        bot.send_message(message.from_user.id, 'Обращение отменено.')
    else:
        problem = 'FEED_BACK\n' + text_problem + message.text
        Other_function.logging_event('info', problem)
        text_message = 'Ваше обращение создано!\n' + 'Тип: ' + text_problem + 'Текст сообщения: ' + message.text
        bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
        bot.send_message(chat_id=Data.list_admins.get('Никита'), text='Поступило новое обращение от пользователя:\n' +
                                                                      'Тип: ' + text_problem + 'Текст сообщения: ' +
                                                                      message.text)


@bot.message_handler(commands=['create_record'])
def create_record(message):
    print(full_name_user(message) + 'отправил команду:\n' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        if SQLite.check_for_admin(message.from_user.id) is True:  # Если пользователь админ
            text_message = 'В какой лист добавить запись?'
            keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            buttons = ['Уведомления для всех', 'Уведомления для подписчиков', 'Уведомления для админов']
            keyboard.add(*buttons)
            bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)  #
            bot.register_next_step_handler(message, create_record_step_2)  # Регистрация следующего действия
            print(answer_bot + text_message + '\n')
        else:  # Если юзер не админ, он получит следующее сообщение
            text_message = 'У вас нет прав для выполнения этой команды'
            bot.send_message(message.chat.id, text_message)
    else:  # Если пользователь не зарегистрирован, он получит следующее сообщение
        text_message = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, text_message)


def create_record_step_2(message):
    print(full_name_user(message) + 'написал:\n' + message.text)
    list_of_answers = [message.text]
    hide_keyboard = telebot.types.ReplyKeyboardRemove()
    text_message = 'Введи текст уведомления'
    bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
    bot.register_next_step_handler(message, create_record_step_3, list_of_answers)  # Регистрация следующего действия


def create_record_step_3(message, list_of_answers):
    print(full_name_user(message) + 'написал:\n' + message.text)
    list_of_answers.append(message.text)
    text_message = 'Введи дату когда уведомить о событии в формате <ДД.ММ.ГГГГ>. \n*Примечание: если дату указать ' \
                   'некорректно, уведомление не сработает!'
    bot.send_message(message.from_user.id, text_message)
    bot.register_next_step_handler(message, create_record_step_4, list_of_answers)  # Регистрация следующего действия


def create_record_step_4(message, list_of_answers):
    print(full_name_user(message) + 'написал:\n' + message.text)
    list_of_answers.append(message.text)
    print(list_of_answers)
    sheet_name = list_of_answers[0]
    text_notification = list_of_answers[1]
    date_notification = list_of_answers[2]
    opener = urllib.request.build_opener(SMBHandler)
    file_name = opener.open(Data.route)
    wb = load_workbook(file_name)  # Открываем нужную книгу
    sheet = wb[sheet_name]  # Получить лист по ключу
    column_a = sheet['A']  # Колонка A
    print(len(column_a))

    i = 0
    while column_a[i]:  # Повторить пока в колонке А не пусто
        if column_a[i].value is None:
            print('Строка ' + str(column_a[i].row) + '-' + str(column_a[i].value) + ' пустая')
            empty_string = column_a[i].row
            sheet.cell(row=empty_string, column=1).value = date_notification  #
            sheet.cell(row=empty_string, column=2).value = text_notification  #
            wb.save('test.xlsx')  # Сохранить книгу
            file = open('test.xlsx', 'rb')
            file_name = opener.open(Data.route, data=file)
            file_name.close()
            os.remove('test.xlsx')
            break

        if i < len(column_a):
            i += 1

    text_message = '• Запись добавлена в лист: "' + str(list_of_answers[0]) + '"\n' + \
                   '• Текст: "' + str(text_notification) + '"\n' + \
                   '• Дата уведомления: "' + str(date_notification) + '"\n'
    Notifications.notification_for(message.from_user.first_name + ' создал новое событие\n\n' + text_message,
                                   'status', 'admin')
    # bot.send_message(chat_id=Data.list_admins.get('Никита'),
    #                  text=message.from_user.first_name + ' создал новое событие\n\n' + text_message)
    list_of_answers.clear()
    exit()


@bot.message_handler(commands=['games'])
def games(message):
    print(full_name_user(message) + 'отправил команду ' + message.text)
    if SQLite.check_for_existence(message.from_user.id) is True:  # Проверка на наличие юзера в БД
        SQLite.update_data_user(message)  # Актуализация данных о пользователе в БД
        text_message = 'На данный момент доступна одна игра'
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        buttons = ['Играть в "Угадаю число"', 'Отмена']
        keyboard.add(*buttons)
        bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)  #
        bot.register_next_step_handler(message, games_step_2)  # Регистрация следующего действия
        print(answer_bot + text_message + '\n')
    else:  # Если пользователь не зарегистрирован, бот предложит это сделать
        end_text = 'Чтобы воспользоваться функцией нужно зарегистрироваться, жми /start'
        bot.send_message(message.from_user.id, end_text)
        print(answer_bot + end_text + '\n')


def games_step_2(message):
    if message.text == 'Играть в "Угадаю число"':
        text_message = 'Хорошо, начнём. Правила просты - загадай число от 1 до 100 а я попробую его угадать. ' \
                       'Я называю число, если твоё число меньше, жми "меньше", если твоё число больше, ' \
                       'жми "больше", а если угадал - "в точку".'
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        buttons = ['Начнём']
        keyboard.add(*buttons)
        bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)
        # bot.send_message(message.from_user.id, text_message)
        bot.register_next_step_handler(message, games_step_3)
        # return int(number)
    elif message.text == 'Отмена':
        text_message = 'Хорошо, сыграем в следующий раз.'
        bot.send_message(message.from_user.id, text_message)
    # else:
    # return int(number)


def games_step_3(message):
    if message.text == 'Начнём':
        number = random.randint(0, 100)
        text_message = 'Возможно это ' + str(number) + ' ?'
        keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        buttons = ['Больше', 'Меньше', 'В точку']
        keyboard.add(*buttons)
        bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)
        bot.register_next_step_handler(message, games_step_4, number, lower=1, high=100, count=1)


def games_step_4(message, number, lower, high, count):
    hide_keyboard = telebot.types.ReplyKeyboardRemove()
    if message.text == 'Больше':
        print(message.text)
        lower = number + 1
    elif message.text == 'Меньше':
        print(message.text)
        high = number - 1
    elif message.text == 'В точку':
        text_message = 'Я угадал твоё число за ' + str(count) + ' ходов'
        bot.send_message(message.from_user.id, text_message, reply_markup=hide_keyboard)
        exit()
    else:
        print(message.text)
    middle = (high + lower) // 2
    text_message = 'Я думаю твоё число ' + str(middle)
    print(text_message)
    count += 1
    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    buttons = ['Больше', 'Меньше', 'В точку']
    keyboard.add(*buttons)
    bot.send_message(message.from_user.id, text_message, reply_markup=keyboard)
    bot.register_next_step_handler(message, games_step_4, middle, lower, high, count)
    print('среднее: ' + str(middle) + '\nнижний предел: ' + str(lower) + '\nверхний предел: ' + str(
        high) + '\nпопытка: ' + str(count) + '\n')


@bot.message_handler(content_types=['text'])
def other_functions(message):
    SQLite.update_data_user(message)
    print(full_name_user(message) + 'написал:\n' + message.text)
    i_can = "Чтобы узнать что я умею напиши /help."
    bot.send_message(message.chat.id, i_can)
    print(answer_bot + i_can + '\n')


if __name__ == '__main__':
    while True:
        try:
            bot.polling(none_stop=True)
        except Exception as e:
            time.sleep(3)
            bot.send_message(chat_id=Data.list_admins.get('Никита'), text='Бот выдал ошибку: ' + str(e))
            print(str(e))
            Other_function.logging_event('error', str(e))
            # os.kill(os.getpid(), 9)
