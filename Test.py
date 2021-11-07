import datetime
import urllib

from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler
import urllib.request

import Data


def read_sheet(sheet, event_number):
    opener = urllib.request.build_opener(SMBHandler)
    file_name = opener.open(Data.route)
    wb = load_workbook(file_name)  # Открываем нужную книгу
    sheet = wb[sheet]  # Получить лист по ключу
    column_a = sheet['A']  # Колонка A
    column_b = sheet['B']  # Колонка B
    column_c = sheet['C']  # Колонка C

    now_date = datetime.datetime.now()  # Получаем текущую дату
    date_list = []
    if column_a and column_b and column_c is not None:
        for i in range(len(column_a)):
            if i is not None:
                if isinstance(column_a[i].value, datetime.datetime):
                    value_one = column_a[i].value
                    value_one_column = column_a[i].column  # Колонка значения 1й даты
                    value_one_row = column_a[i].row  # Строка значения 1й даты
                    value_two = sheet.cell(row=value_one_row, column=(value_one_column + 1)).value
                    difference_date = value_one - now_date
                    difference_date = difference_date.days + 1
                    if isinstance(value_two, datetime.datetime):
                        meaning = sheet.cell(row=value_one_row, column=(value_one_column + 2)).value
                    else:
                        meaning = value_two

                    if difference_date >= 0:
                        if isinstance(value_two, datetime.datetime):
                            date_list.append([value_one, value_two, meaning])
                        else:
                            date_list.append([value_one, meaning])
            else:
                print('Данные отсутствуют')

    date_list.sort()

    if len(date_list[event_number - 1]) == 2:
        date = date_list[(event_number - 1)][0]
        date = date.strftime("%d.%m.%Y")
        event = date_list[(event_number - 1)][1]
        event_data = [date, event]
    elif len(date_list[event_number - 1]) == 3:
        first_date = date_list[(event_number - 1)][0]
        first_date = first_date.strftime("%d.%m.%Y")
        last_date = date_list[(event_number - 1)][1]
        last_date = last_date.strftime("%d.%m.%Y")
        event = date_list[(event_number - 1)][2]
        event_data = [first_date, last_date, event]

    print(event_data)

    return event_data


read_sheet('Дежурный', 1)
# read_sheet('Инвентаризация', 8)
