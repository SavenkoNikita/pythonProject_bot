import datetime
import logging
import os
import random
import sqlite3
import time
import traceback
import urllib
from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler
from telebot import types
import requests
import urllib.request
from requests.auth import HTTPBasicAuth
import xml.etree.ElementTree as ET

import Data
from Data import list_command_admin, list_command_user


def get_key(user_name):
    """Проверяет среди словаря есть ли в нём имя{user_name} и возвращает соответствующий id_telegram"""

    # key = 'Имя не найдено или человек не относится к дежурным. Error Other_function.get_key.'
    for keys, v in Data.user_data.items():
        for a in v:
            if a == user_name:
                key = keys
                return key


def get_data_user_SQL(user_name):
    """Принимает на вход Имя, проверяет есть ли это имя в списке дежурных, если есть, получает user_id, если он есть в
    БД, подтягивает Имя и user_name. Результат - <имя + @username>"""

    id_telegram = get_key(user_name)  # Присваиваем id из get_key
    print(id_telegram)
    if id_telegram is not None:  # Если id есть то
        if SQL().check_for_existence(id_telegram) is True:  # Если в SQL есть такой id
            end_text = SQL().get_user_info(id_telegram)  # Получаем склейку <имя + @username>
            print(end_text)
            return end_text


def logging_event(condition, text):
    """Возможные варианты записи логов - 'debug', 'info', 'warning', 'error', 'critical'"""
    if text is not None:
        logging.basicConfig(filename=Data.way_to_log_file, level=logging.INFO,
                            format="%(asctime)s - [%(levelname)s] - %(message)s")
        if condition == 'debug':
            logging.debug(text)
        elif condition == 'info':
            logging.info(text)
        elif condition == 'warning':
            logging.warning(text)
        elif condition == 'error':
            logging.error(text)
        elif condition == 'critical':
            logging.critical(text)


def can_do_it(x):
    """Перечисляет строка за строкой всё что есть в списке с переводом строки."""

    cd = ('\n'.join(map(str, x)))
    return cd


def can_help(user_id):
    """Формирует список доступных команд для пользователя в зависимости админ он или нет."""

    end_text = 'Вот что я умею:' + '\n'
    check_admin = SQL().check_for_admin(user_id)
    if check_admin is True:  # Если пользователь админ
        end_text = end_text + can_do_it(list_command_admin)  # Передать полный список доступных команд
    else:  # Если пользователь НЕ админ
        end_text = end_text + can_do_it(list_command_user)  # Передать список команд доступных юзеру
    return end_text


def random_name():
    """Присылает уведомление кто сегодня закрывает сигналы"""
    if datetime.datetime.today().isoweekday() <= 5:
        list_name = ['Паша', 'Дима', 'Никита']
        random_name = random.choice(list_name)
        end_text = f'Случайным образом определено, что сигналы сегодня закрывает {random_name}'
        for i in list_name:
            Data.bot.send_message(chat_id=Data.list_admins.get(i), text=end_text)


class SQL:
    """Проверка, добавление, обновление и удаление данных о пользователях"""

    def __init__(self):
        self.sqlite_connection = sqlite3.connect(Data.way_sql, check_same_thread=False, timeout=10)
        self.cursor = self.sqlite_connection.cursor()

    def check_for_existence(self, user_id, table_DB='users'):
        """Проверка на существование пользователя в БД"""

        try:
            info = self.cursor.execute(f'SELECT * FROM {table_DB} WHERE user_id=?', (user_id,))
            if info.fetchone() is None:  # Если человека нет в бд
                return False
            else:  # Если есть человек в бд
                return True
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))

    def check_for_admin(self, user_id):
        """Проверка на то, является ли пользователь админом"""
        if self.check_for_existence(user_id) is True:
            info = self.cursor.execute('SELECT * FROM users WHERE status=? and user_id=?', ('admin', user_id))
            if info.fetchone() is None:  # Если пользователь не админ
                return False
            else:  # Если пользователь админ
                return True

    # def check_status_bd(self, user_id, column_bd):
    #     """Проверяет статус колонки в БД. Возвращает булево."""
    #     if self.check_for_existence(user_id) is True:
    #         info = self.cursor.execute(f'SELECT * FROM users WHERE {column_bd}=? and user_id=?', ('yes', user_id))
    #         if info.fetchone() is None:  # Если пользователь НЕ подписан на рассылку
    #             return False
    #         else:  # Если пользователь подписан на рассылку
    #             return True

    def change_status_DB(self, user_id, column_bd):
        """Меняет у {user_id} в БД статус {column_bd} c 'yes' на 'no' и наоборот в зависимости от текущего статуса."""

        status = self.check_status_DB(user_id, column_bd, 'yes')

        if status is False:
            self.update_sqlite_table('yes', user_id, column_bd)
        else:
            self.update_sqlite_table('no', user_id, column_bd)

    # def set_unsubscribe(self, user_id):
    #     """Присваивает статус <отписан> в БД"""
    #     self.update_sqlite_table('no', user_id, 'notification')

    def db_table_val(self, user_id, first_name, last_name, username):
        """Проверка на уникальность и добавление данных о пользователе в SQL"""
        # Словарь вносимых в базу данных значений
        list_user = {
            'ID: ': user_id,
            'Имя: ': first_name,
            'Фамилия: ': last_name,
            'Username:  @': username
        }

        def data_user():  # Заполняем отсутствующие о пользователе данные (имя, ник и тп) фразой <нет данных>
            for a, b in list_user.items():
                if b is None:
                    b = '<нет данных>'
                    return str(a) + str(b) + '\n'
                else:
                    return str(a) + str(b) + '\n'

        if self.check_for_existence(user_id) is False:
            sqlite_select_query = 'SELECT * from users where user_id'
            self.cursor.execute(sqlite_select_query)
            records = self.cursor.fetchall()
            print('ID всех пользователей:\n')
            all_user_sql = []
            for row in records:
                all_user_sql.append(row[1])
            text_message = 'Присоединился новый пользователь. Нас уже ' + str(len(all_user_sql) + 1) + '!'
            logging_event('info', str(text_message))

            i = 0
            while i < len(all_user_sql):
                user_name = SQL().get_user_info(all_user_sql[i])
                try:
                    print(user_name)
                    Data.bot.send_message(all_user_sql[i], text=text_message)
                    # Data.bot.send_message(Data.list_admins.get('Никита'), text=text_message)
                except Data.telebot.apihelper.ApiTelegramException:
                    text_error = 'Пользователь <' + user_name + '> заблокировал бота!'
                    print(text_error)
                    logging_event('error', str(text_error))
                    self.log_out(all_user_sql[i])  # SQL().log_out(all_user_sql[i])
                i += 1
            self.cursor.execute('INSERT INTO users (user_id, user_first_name, user_last_name, username) VALUES (?, ?, '
                                '?, ?)',
                                (user_id, first_name, last_name, username))
            self.sqlite_connection.commit()
            self.cursor.close()
            end_text = 'К боту подключился новый пользователь!\n' + data_user() + '\n'
            Data.bot.send_message(chat_id=Data.list_admins.get('Никита'), text=end_text)
            print(end_text)
        elif self.check_for_existence(user_id) is True:
            # обновление изменений данных о пользователе:
            sqlite_update_query = f'UPDATE users set user_first_name = ?, user_last_name = ?, username = ? ' \
                                  f'WHERE user_id = {user_id}'
            column_values = (first_name, last_name, username)
            self.cursor.execute(sqlite_update_query, column_values)
            self.sqlite_connection.commit()
            self.cursor.close()
            end_text = 'Обновлены данные пользователя\n' + data_user() + '\n'
            Data.bot.send_message(chat_id=Data.list_admins.get('Никита'), text=end_text)
            print(end_text)
        else:
            end_text = 'Пользователь уже есть в базе данных!\n' + data_user() + '\n'
            Data.bot.send_message(chat_id=Data.list_admins.get('Никита'), text=end_text)
            print(end_text)

        return end_text

    def update_sqlite_table(self, status, user_id, column_name):
        """Обновление статуса пользователя в SQL"""
        if self.check_for_existence(user_id) is True:
            try:
                self.cursor.execute(f"Update users set {column_name} = ? where user_id = ?", (status, user_id))
                self.sqlite_connection.commit()
                print("Запись успешно обновлена")
                self.cursor.close()
            except sqlite3.Error as error:
                print("Ошибка при работе с SQLite", error)
                logging_event('error', str(error))
            finally:
                if self.sqlite_connection:
                    self.sqlite_connection.close()

    def log_out(self, user_id, table_name_DB='users'):
        """Стереть все данные о пользователе из БД"""
        try:
            if self.sqlite_connection:
                try_message = f'Все данные о пользователе <{self.get_user_info(user_id)}> успешно удалены из {table_name_DB}! '
                print(try_message)
                logging_event('info', try_message)
                self.cursor.execute(f'DELETE from {table_name_DB} where user_id = ?', (user_id,))
                self.sqlite_connection.commit()
                self.cursor.close()
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))

    def update_data_user(self, message):
        """Обновить данные о пользователе"""
        user_id = message.from_user.id
        # user_id = message.id и далее везде
        first_name = message.from_user.first_name
        last_name = message.from_user.last_name
        username = message.from_user.username
        if self.check_for_existence(user_id) is True:
            try:
                self.cursor.execute('SELECT * FROM users WHERE user_id=?', (user_id,))
                records = self.cursor.fetchall()
                for row in records:
                    user_id_SQL = row[1]
                    user_first_name_SQL = row[2]
                    user_last_name_SQL = row[3]
                    username_SQL = row[4]

                    if user_id == user_id_SQL or \
                            first_name != user_first_name_SQL or \
                            last_name != user_last_name_SQL or \
                            username != username_SQL:
                        self.db_table_val(user_id, first_name, last_name, username)
                self.cursor.close()
            except sqlite3.Error as error:
                print("Ошибка при работе с SQLite: ", error)
                logging_event('error', str(error))
            # finally:
            #     if self.sqlite_connection:
            #         self.sqlite_connection.close()

    def get_user_info(self, user_id):
        """Получить данные о пользователе"""
        try:
            self.cursor.execute("""select * from users where user_id = ?""", (user_id,))
            records = self.cursor.fetchall()
            for row in records:
                if row[4] is not None:  # Если в SQL есть запись о username
                    name_and_username = f'{row[2]} @{row[4]}'  # Получаем имя и username
                else:
                    name_and_username = row[2]  # Получаем имя
                return name_and_username
            self.cursor.close()
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        # finally:
        #     if self.sqlite_connection:
        #         self.sqlite_connection.close()

    def get_user_sticker(self, user_id):
        """Получить стикер пользователя"""
        try:
            sql_select_query = """select * from users where user_id = ?"""
            self.cursor.execute(sql_select_query, (user_id,))
            records = self.cursor.fetchall()
            for row in records:
                if row[7] is not None:  # Если в SQL есть запись о
                    return row[7]  # Получаем
                else:
                    return None  # Получаем
            self.cursor.close()
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        # finally:
        #     if self.sqlite_connection:
        #         self.sqlite_connection.close()

    def get_list_users(self):
        """Получить список всех пользователей"""
        try:
            sql_select_query = """select * from users"""
            self.cursor.execute(sql_select_query)
            records = self.cursor.fetchall()
            count_records = 'Общее количество пользователей: ' + str(len(records)) + '\n\n'
            list_data = [count_records]
            for row in records:
                user_data = 'Имя: ' + str(row[2]) + \
                            '\nФамилия: ' + str(row[3]) + \
                            '\nСтатус: ' + str(row[5]) + \
                            '\nПодписка: ' + str(row[6]) + '\n\n'
                list_data.append(user_data)
            full_list = ''.join(list_data)
            print(full_list)
            return full_list
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        # finally:
        #     if self.sqlite_connection:
        #         self.sqlite_connection.close()

    def set_user(self, user_id):
        """Установить пользователю права юзера"""
        self.update_sqlite_table('user', user_id, 'status')

    def set_admin(self, user_id):
        """Установить пользователю права админа"""
        self.update_sqlite_table('admin', user_id, 'status')

    def check_status_DB(self, user_id, column_name, status, table_DB='users'):
        f"""Проверяет у конкретного {user_id} соответствует ли {status} в колонке {column_name}."""

        try:
            request = f'select * from {table_DB} where user_id = ? and {column_name} = ?'
            self.cursor.execute(request, (user_id, status))
            user = self.cursor.fetchone()
            if user is not None:
                return True
            else:
                return False
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))

    def create_list_users(self, column_name_DB, column_meaning, name_table_DB='users'):
        """Формирует список пользователей с подходящими параметрами. Например, список id_user которые подписаны на
        уведомления. На вход принимает имя колонки из БД и значение ячейки. Результат list[user_id]."""

        try:
            info = self.cursor.execute(f'SELECT * FROM {name_table_DB} WHERE {column_name_DB}=?', (column_meaning,))
            records = self.cursor.fetchall()

            list_users_id = []

            for row in records:
                if info.fetchone() is None:  # Если есть совпадение по запросу
                    list_users_id.append(row[1])

            if len(list_users_id) == 0:
                return None
            else:
                return list_users_id
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def get_mess_id(self, user_id):
        """Проверяет есть ли message_id в tracking_sensor_defroster и если есть возвращает его"""

        try:
            request = f'select * from tracking_sensor_defroster where user_id = ?'
            self.cursor.execute(request, (user_id,))
            row = self.cursor.fetchone()
            if row is not None:
                return row[1]
            else:
                return None
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def add_user_by_table(self, user_id, column_name, status, name_table_DB):
        f"""Если пользователь подписан на {column_name} и отсутствует в {name_table_DB}, добавляет его."""

        try:
            if self.check_status_DB(user_id, column_name, status) is True:
                if self.check_for_existence(user_id, name_table_DB) is False:
                    self.cursor.execute(f'INSERT INTO {name_table_DB} (user_id, message_id) VALUES (?, ?)',
                                        (user_id, None))
                    self.sqlite_connection.commit()
                    self.cursor.close()
                    print(f'Пользователь {user_id} успешно добавлен в таблицу {name_table_DB}')
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def update_mess_id_by_table(self, user_id, message_id, table_name_DB, name_column_in_table_users):
        f"""В таблице {table_name_DB} обновляет пользователю {user_id} значение {message_id}"""

        try:
            if self.check_status_DB(user_id, name_column_in_table_users, 'yes') is True:
                if self.check_for_existence(user_id, table_name_DB) is True:
                    # обновление изменений данных о сообщении:
                    sqlite_update_query = f'UPDATE {table_name_DB} set message_id={message_id} ' \
                                          f'WHERE user_id={user_id}'
                    self.cursor.execute(sqlite_update_query)
                    self.sqlite_connection.commit()
                    self.cursor.close()
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def get_dict(self, table_name_DB):
        """Получить список user_id и message_id"""

        try:
            request = f'select * from {table_name_DB}'
            self.cursor.execute(request)
            row = self.cursor.fetchall()
            return row
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def talk(self, text_message):
        """Ищет в БД ответ на вопрос. Принимает текст сообщения. Если есть ответ, возвращает его,
        а если нет, возвращает None"""

        try:
            request = f'select * from talk where question = ?'
            self.cursor.execute(request, (text_message,))
            row = self.cursor.fetchone()
            if row is not None:
                return row[1]
            else:
                self.insert_data_speak_DB(text_message)
                return None
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def insert_data_speak_DB(self, text_message):
        """Если вопрос отсутствует в БД, добавляет его."""

        try:
            self.cursor.execute(f'INSERT INTO talk (question, answer) VALUES (?, ?)',
                                (text_message, None))
            self.sqlite_connection.commit()
            self.cursor.close()
            print(f'Запись {text_message} успешно добавлена в таблицу talk')
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def update_answer_speak_DB(self, question, answer):
        """"""

        try:
            sqlite_update_query = f'UPDATE talk set answer=? WHERE question=?'
            self.cursor.execute(sqlite_update_query, (answer, question,))
            self.sqlite_connection.commit()
            self.cursor.close()
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def count_not_answer(self, count=0):
        """Считает кол-во вопросов без ответа. Результат - число."""

        try:
            self.cursor.execute('select * from talk')
            row = self.cursor.fetchall()
            for rows in row:
                # print(rows[1])
                if rows[1] is None:
                    count += 1

            # count_none_rows = f'На данный момент есть {count} вопросов без ответа.'
            return count
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    def search_not_answer(self):
        """Находит первый вопрос без ответа. Результат - текст вопроса."""

        try:
            self.cursor.execute('select * from talk')
            row = self.cursor.fetchall()
            for rows in row:
                if rows[0] is not None:
                    if rows[1] is None:
                        text = rows[0]
                        return text

            text = 'Нет вопросов без ответа :)'
            return text

        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()


class Notification:
    """Методы уведомлений"""

    def __init__(self):
        self.sqlite_connection = sqlite3.connect(Data.way_sql)
        self.cursor = self.sqlite_connection.cursor()

    def all_registered(self, text_message):
        """Функция для уведомления всех пользователей находящихся в БД"""
        try:
            self.cursor.execute('SELECT * from users where user_id')
            records = self.cursor.fetchall()
            all_user_sql = []
            for row in records:
                all_user_sql.append(row[1])
            self.cursor.close()
            i = 0
            print('Уведомление отправлено следующим пользователям:\n')
            while i < len(all_user_sql):
                username = SQL().get_user_info(all_user_sql[i])
                try:
                    print(username)
                    Data.bot.send_message(all_user_sql[i], text=text_message)
                    # Data.bot.send_message(Data.list_admins.get('Никита'), text=text_message)
                except Data.telebot.apihelper.ApiTelegramException:
                    text_error = 'Пользователь <' + username + '> заблокировал бота!'
                    print(text_error)
                    logging_event('error', str(text_error))
                    SQL().log_out(all_user_sql[i])
                i += 1
        except sqlite3.Error as error:
            print('Ошибка при работе с SQLite', error)
            logging_event('error', error)
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()
                print('Соединение с SQLite закрыто')

    def notification_for(self, text_message, column, column_meaning):
        f"""Уведомления для юзеров с указанными параметрами. Принимает {text_message}, название столбца в БД {column}, 
        и искомое значение этой колонки {column_meaning}. Например - я хочу уведомить подписчиков о событии, 
        это будет выглядеть так: notification_for('Сообщение, которое хочу передать', 'notification', 'yes')."""

        try:
            self.cursor.execute(f'SELECT * FROM users WHERE {column}= ?', [column_meaning])
            records = self.cursor.fetchall()
            # print('Список ID:\n')
            all_id_sql = []
            for row in records:
                all_id_sql.append(row[1])
            self.cursor.close()
            # print(all_id_sql)
            i = 0
            print('Уведомление отправлено следующим пользователям:\n')
            while i < len(all_id_sql):
                username = SQL().get_user_info(all_id_sql[i])
                try:
                    print(username)
                    Data.bot.send_message(all_id_sql[i], text=text_message)
                    # Data.bot.send_message(Data.list_admins.get('Никита'), text=text_message)
                except Data.telebot.apihelper.ApiTelegramException:
                    text_error = 'Пользователь <' + username + '> заблокировал бота!'
                    print(text_error)
                    logging_event('error', str(text_error))
                    SQL().log_out(all_id_sql[i])
                i += 1
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', error)
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()
                print("Соединение с SQLite закрыто")

    def send_sticker_for(self, user_first_name, column, column_meaning):
        """Уведомления для юзеров с указанными параметрами"""
        try:
            self.cursor.execute(('SELECT * FROM users WHERE ' + column + ' = ?'), [column_meaning])
            records = self.cursor.fetchall()
            # print('Список ID:\n')
            all_id_sql = []
            for row in records:
                all_id_sql.append(row[1])
            self.cursor.close()
            # print(all_id_sql)
            i = 0
            print('Стикер отправлен следующим пользователям:\n')
            while i < len(all_id_sql):
                username = SQL().get_user_info(all_id_sql[i])
                try:
                    print(username)
                    user_sticker = SQL().get_user_sticker(get_key(user_first_name))
                    Data.bot.send_sticker(all_id_sql[i], user_sticker)
                    # Data.bot.send_sticker(Data.list_admins.get('Никита'), user_sticker)
                except Data.telebot.apihelper.ApiTelegramException:
                    print('Пользователь <' + username + '> заблокировал бота!')
                    SQL().log_out(username)
                except Exception as e:
                    time.sleep(3)
                    Data.bot.send_message(chat_id=Data.list_admins.get('Никита'), text='Бот выдал ошибку: ' + str(e))
                    print(str(e))
                    logging_event('error', str(e))
                i += 1
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()
                print("Соединение с SQLite закрыто")

    def repeat_for_list(self, data_list, user_id, count=1):
        """Присылает {count} элементов из списка {data_list} пользователю {user_id}. Если не указать {count},
        пришлёт первый элемент из списка."""

        temporary_list = []
        for elem in range(0, count):
            Data.bot.send_message(user_id, data_list[elem])
            temporary_list.append(data_list[elem])

        print('Бот ответил:')
        for i in temporary_list:
            print(i)

    def update_mess(self, name_table_DB, func):
        """Обновляет сообщение у пользователей."""

        try:
            text_message = func
            data_list = SQL().get_dict(name_table_DB)
            if len(data_list) != 0:
                for elem in data_list:
                    user_id = elem[0]
                    message_id = elem[1]
                    Data.bot.edit_message_text(text=text_message, chat_id=user_id, message_id=message_id,
                                               parse_mode='Markdown')
        except sqlite3.Error as error:
            print("Ошибка при работе с SQLite", error)
            logging_event('error', str(error))
        finally:
            if self.sqlite_connection:
                self.sqlite_connection.close()

    # def update_mess_all_sensor(self):
    #     """"""
    #
    #     try:
    #         text_message = Tracking_sensor.TrackingSensor().check_all()


class File_processing:
    """Обработка файла"""

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.opener = urllib.request.build_opener(SMBHandler)
        self.file_name = self.opener.open(Data.route)
        self.wb = load_workbook(self.file_name)  # Открываем нужную книгу
        self.sheet = self.wb[sheet_name]  # Получить лист по имени
        self.first_column = self.sheet['A']
        self.now_date = datetime.datetime.now()  # Получаем текущую дату
        self.title_first_column = self.sheet.cell(row=1, column=1).value  # Заголовок первой колонки
        self.title_second_column = self.sheet.cell(row=1, column=2).value  # Заголовок второй колонки
        self.title_third_column = self.sheet.cell(row=1, column=3).value  # Заголовок третьей колонки
        if (self.title_first_column and self.title_second_column and self.title_third_column) is not None:
            """Если заполнены 3 колонки, {self.count_meaning} возвращает '3', а если 2, то '2'"""
            self.count_meaning = 3  # Количество непустых колонок 3
        else:
            self.count_meaning = 2  # Количество непустых колонок 2

    def difference_date(self, date):
        f"""На вход принимает дату. Возвращает int(разницу) в днях между сегодня и {date}. 
        Где "0" означает сегодня, "1" - "завтра" и тд"""

        difference = date - self.now_date  # Разница между 1‑й датой и сегодня
        difference = difference.days + 1  # Форматируем в кол-во дней +1

        if difference == -1:
            print('Событие было вчера')
        elif difference == 0:
            print('Событие сегодня')
        elif difference == 1:
            print('Событие завтра')

        return difference

    def read_file(self):
        """Возвращает списки значений в формате [ДД.ММ.ГГГГ(datetime), ДД.ММ.ГГГГ(datetime), событие].
        Если дата в списке одна, то соответственно [ДД.ММ.ГГГГ(datetime), событие]."""

        self.clear_old_data()

        date_list = []  # Объявляем пустой список

        try:
            for i in range(2, 100):  # Повторить для каждого значения в колонке А
                if self.sheet.cell(row=i, column=1).value is not None:  # Если ячейка не пустая
                    if self.count_meaning == 2:  # Если заполнены 2 колонки
                        date = self.sheet.cell(row=i, column=1).value  # Ячейка с датой
                        if isinstance(date, datetime.datetime):  # Если значение == дата
                            date = date
                        else:
                            date = datetime.datetime.strptime(date, '%d.%m.%Y')
                        meaning = self.sheet.cell(row=i, column=2).value  # Ячейка с событием
                        line = [date, meaning]  # Лист из 2 значений
                        date_list.append(line)  # Добавить лист в общий список
                    elif self.count_meaning == 3:  # Если заполнены 3 колонки
                        date_one = self.sheet.cell(row=i, column=1).value  # Ячейка с первой датой
                        if isinstance(date_one, datetime.datetime):  # Если значение == дата
                            date_one = date_one
                        else:
                            date_one = datetime.datetime.strptime(date_one, '%d.%m.%Y')

                        date_two = self.sheet.cell(row=i, column=2).value  # Ячейка со второй датой
                        if isinstance(date_two, datetime.datetime):  # Если значение == дата
                            date_two = date_two
                        else:
                            date_two = datetime.datetime.strptime(date_two, '%d.%m.%Y')

                        meaning = self.sheet.cell(row=i, column=3).value  # Ячейка с событием
                        line = [date_one, date_two, meaning]  # Лист из 3 значений
                        date_list.append(line)  # Добавить лист в общий список
        except Exception:
            print(traceback.format_exc())
        finally:
            date_list.sort()  # Сортируем лист по дате

            if len(date_list) == 0:  # Если список пуст
                print(f'Список <{self.sheet_name}> пуст\n')
                return None
            else:
                return date_list

    def clear_old_data(self):
        """Очистка неактуальных данных"""

        for i in range(2, 10):  # Повторить для каждого значения в 1 колонке
            date = self.sheet.cell(row=i, column=1).value
            # print(date)
            # print(type(date))
            if date is not None:  # Если значение не пустое
                if isinstance(date, datetime.datetime):  # Если значение == дата
                    # print('date')
                    date = date
                else:
                    date = datetime.datetime.strptime(date, '%d.%m.%Y')
                    # print(f'{date} not date')

                if self.difference_date(date) < 0:  # Если событие в прошлом
                    date_event = self.sheet.cell(row=i, column=1)  # Колонка с датой
                    event = self.sheet.cell(row=i, column=self.count_meaning).value  # Ячейка с событием
                    print(f'Удалена строка {date_event.row}\n'
                          f'Дата: {date_event.value}\n'
                          f'Текст: {event}\n\n')
                    self.sheet.delete_rows(i)  # Удаляем указанную в скобках строку
                    self.wb.save('test.xlsx')  # Сохранить книгу
                    file = open('test.xlsx', 'rb')
                    self.file_name = self.opener.open(Data.route, data=file)
                    self.file_name.close()
                    os.remove('test.xlsx')
                    time.sleep(1)
                    self.clear_old_data()
                    # else:
                    #     print(f'{self.difference_date(self.sheet.cell(row=i, column=1).value)}')
                # else:
                #     print(f'{type(self.sheet.cell(row=i, column=1).value)} не дата')
            # else:
            #     print(f'{self.sheet.cell(row=i, column=1).value} is None')

    def next_dej(self):
        """Если файл заполнен, возвращает строку 'В период с {first_date} по {second_date} будет дежурить {name}.'"""

        data_list = self.read_file()

        if self.sheet_name == 'Дежурный':
            self.clear_old_data()
            if data_list is not None:
                first_date = data_list[0][0].strftime('%d.%m.%Y')  # Дата str(1)
                second_date = data_list[0][1].strftime('%d.%m.%Y')  # Дата str(2)
                name_from_SQL = SQL().get_user_info(get_key(self.read_file()[0][2]))  # Имя дежурного

                text_message = f'В период с {first_date} по {second_date} будет дежурить {name_from_SQL}.'

                return text_message

    def sticker_next_dej(self):
        """Сравнивает имя из файла с БД, и если есть совпадение и в БД содержится стикер то вернёт его."""

        self.clear_old_data()
        if self.read_file() is not None:
            sticker = SQL().get_user_sticker(get_key(self.sheet.cell(row=2, column=self.count_meaning).value))
            if sticker is not None:
                return sticker
            else:
                return None

    def list_dej(self):
        """Возвращает список всех дежурных, что есть в списке"""

        list_dej = []
        for i in self.read_file():
            if self.sheet_name == 'Дежурный':
                if i is not None:
                    date_one = i[0].strftime("%d.%m.%Y")
                    date_two = i[1].strftime("%d.%m.%Y")
                    name_from_SQL = SQL().get_user_info(get_key(i[2]))
                    text_message = f'В период с {date_one} по {date_two} будет дежурить {name_from_SQL}.'
                    list_dej.append(text_message)

        return list_dej

    def next_invent(self):
        """Если файл заполнен, возвращает строку 'До предстоящей инвентаризации осталось {N} дней.
        Судя по графику, выходит {name}'"""

        self.clear_old_data()

        if self.sheet_name == 'Инвентаризация':
            if self.read_file() is not None:
                name_from_SQL = SQL().get_user_info(get_key(self.sheet.cell(row=2, column=2).value))
                if name_from_SQL is None:
                    name_from_SQL = self.sheet.cell(row=2, column=2).value
                dead_line_date = self.difference_date(self.sheet.cell(row=2, column=1).value)
                dead_line_text = Counter().days_before_inventory(dead_line_date)
                text_who = f'Судя по графику, выходит {name_from_SQL}.'  # Имя следующего дежурного
                text_message = f'{dead_line_text} {text_who}'
                return text_message
            elif self.read_file() is None:
                text_message = 'Нет данных об этом. Необходимо заполнить файл!'
                return text_message

    def check_event_today(self):
        """Проверяет есть ли сегодня событие. Результат - уведомление соответствующей группе пользователей."""

        data_list = self.read_file()

        if data_list is not None:
            if self.sheet_name in Data.sheets_file:
                for i in data_list:
                    date = i[0]
                    meaning = i[1]
                    if len(i) == 2:
                        if self.difference_date(date) == 0:
                            if self.sheet_name == 'Уведомления для всех':
                                text_message = f'• Уведомление для зарегистрированных пользователей •\n\n' \
                                               f'{meaning}'
                                print(text_message)
                                Notification().all_registered(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для подписчиков':
                                text_message = f'• Уведомление для подписчиков •\n\n' \
                                               f'{meaning}'
                                print(text_message)
                                Notification().notification_for(text_message, 'notification', 'yes')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для админов':
                                text_message = f'• Уведомление для администраторов •\n\n' \
                                               f'{meaning}'
                                print(text_message)
                                Notification().notification_for(text_message, 'status', 'admin')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                        elif 0 <= self.difference_date(date) <= 2:
                            if self.sheet_name == 'Инвентаризация':
                                text_message = f'• Уведомление для администраторов •\n\n' \
                                               f'{self.next_invent()}'
                                print(text_message)
                                Notification().notification_for(text_message, 'status', 'admin')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)

                        time.sleep(5)

    def check_dej_tomorrow(self):
        """Проверяет есть ли завтра дежурный. Результат - уведомление соответствующей группе пользователей."""

        data_list = self.read_file()

        if data_list is not None:
            for i in data_list:
                date = i[0]
                if self.difference_date(date) == 1:  # Если завтра
                    first_date = i[0].strftime('%d.%m.%Y')  # Дата str(1)
                    second_date = i[1].strftime('%d.%m.%Y')  # Дата str(2)
                    name_from_SQL = SQL().get_user_info(get_key(i[2]))  # Имя дежурного

                    text_message = f'В период с {first_date} по {second_date} будет дежурить {name_from_SQL}.'
                    print(datetime.datetime.now(), text_message)

                    sticker_dej = i[2]

                    Notification().notification_for(text_message, 'notification', 'yes')
                    Notification().send_sticker_for(sticker_dej, 'notification', 'yes')
                    # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                elif self.difference_date(date) < 0:  # Если событие в прошлом
                    self.clear_old_data()
                # else:
                #     print(f'Дата {date} не наступила')

    def create_event(self, date, text_event):
        """Принимает дату и текст уведомления и записывает в файл. Возвращает текст с описанием созданного
        уведомления """

        try:
            for i in range(1, 1000):  # Повторить для каждого значения в колонке А
                if self.sheet.cell(row=i, column=1).value is None:
                    print(f'Строка {self.sheet.cell(row=i, column=1).row} пустая')
                    empty_string = self.sheet.cell(row=i, column=1).row
                    date_obj = datetime.datetime.strptime(date, '%d.%m.%Y')
                    date_obj = date_obj.date().strftime('%d.%m.%Y')
                    self.sheet.cell(row=empty_string, column=1).value = date_obj  # Дата
                    self.sheet.cell(row=empty_string, column=2).value = text_event  # Текст события
                    self.wb.save('test.xlsx')  # Сохранить книгу
                    file = open('test.xlsx', 'rb')
                    self.file_name = self.opener.open(Data.route, data=file)
                    self.file_name.close()
                    os.remove('test.xlsx')
                    text_message = f'• Запись добавлена в лист: <{self.sheet_name}>\n' \
                                   f'• Дата уведомления: <{date_obj}>\n' \
                                   f'• Текст: <{text_event}>\n'

                    return text_message
        except urllib as e:
            print(f'Ошибка при работе с файлом:\n{e}')
        except Exception as e:
            time.sleep(3)
            print(str(e))
            logging_event('error', str(e))


class Counter:
    """Класс считает различные сущности"""

    # def __init__(self):

    def days_before_inventory(self, number):
        """Принимает на вход число и склоняет его.
        Результат - str<До предстоящей инвентаризации остался/осталось день/дня/дней>"""

        stayed = ['остался', 'осталось']
        days = ['день', 'дня', 'дней']

        if number == 0:
            return 'Сегодня инвентаризация.'
        elif number % 10 == 1 and number % 100 != 11:
            s = 0
            d = 0
            return 'До предстоящей инвентаризации ' + stayed[s] + ' ' + str(number) + ' ' + days[d] + '.'
        elif 2 <= number % 10 <= 4 and (number % 100 < 10 or number % 100 >= 20):
            s = 1
            d = 1
            return 'До предстоящей инвентаризации ' + stayed[s] + ' ' + str(number) + ' ' + days[d] + '.'
        else:
            s = 1
            d = 2
            return 'До предстоящей инвентаризации ' + stayed[s] + ' ' + str(number) + ' ' + days[d] + '.'

    def number_of_events(self, number):
        """Принимает на вход число и склоняет его.
        Результат - str<На данный момент есть {number} запись/записи/записей. Сколько событий показать?>"""

        records = ['запись', 'записи', 'записей']

        if number == 0:
            return 'К сожалению, нет данных о предстоящих дежурствах.'
        elif number % 10 == 1 and number % 100 != 11:  # <1, 21 запись>
            r = 0
            return 'На данный момент есть ' + str(number) + ' ' + records[r] + '. Сколько событий показать?'
        elif 2 <= number % 10 <= 4 and (number % 100 < 10 or number % 100 >= 20):  # <3, 22 записи>
            r = 1
            return 'На данный момент есть ' + str(number) + ' ' + records[r] + '. Сколько событий показать?'
        else:  # <5, 47 записей>
            r = 2
            return 'На данный момент есть ' + str(number) + ' ' + records[r] + '. Сколько событий показать?'

    def declension_day(self, number):
        """Принимает на вход число и склоняет слово день. Результат - str<день/дня/дней>"""

        days = ['день', 'дня', 'дней']

        if number == 0:
            return days[2]
        elif number % 10 == 1 and number % 100 != 11:
            return days[0]
        elif 2 <= number % 10 <= 4 and (number % 100 < 10 or number % 100 >= 20):
            return days[1]
        else:
            return days[2]


class Bot_menu:
    """Меню бота"""

    def __init__(self):
        self.markup = types.InlineKeyboardMarkup()

        self.list_main_menu = [
            ['Основные функции', 'button_main_functions'],
            ['Управление подписками', 'button_managing_subscriptions'],
            ['Настройка аккаунта', 'button_account_management'],
            ['Дополнительные функции', 'button_additional_functions']
        ]

        self.list_main_functions = [
            ['Узнать кто дежурный', 'button_dej'],
            ['Инвентаризация', 'button_invent'],
            ['Дать пользователю права админа', 'button_admin'],
            ['Лишить пользователя прав админа', 'button_user'],
            ['Создать уведомление', 'button_create_notif']
        ]

        self.list_managing_subscriptions = [
            ['Подписаться на рассылку уведомлений', 'button_subscribe'],
            ['Отказаться от рассылки уведомлений', 'button_unsubscribe'],
            ['Мониторинг дефростеров', 'button_defrosters'],
            ['Мониторинг неисправных датчиков', 'button_all_sensor']
        ]

        self.list_account_management = [
            ['Покинуть чат', 'button_log_out'],
            ['Сменить стикер аккаунта', 'button_sticker']
        ]

        self.list_additional_functions = [
            ['Обратная связь', 'button_feed_back'],
            ['Игры', 'button_games'],
            ['Получить случайное имя', 'button_random'],
            ['Получить список всех пользователей', 'button_all_users']
        ]

    def init_menu(self, data_list_buttons):
        temporary_list = []
        for i in data_list_buttons:
            temporary_list.append(types.InlineKeyboardButton(text=i[0], callback_data=i[1]))

        for elem in temporary_list:
            self.markup.add(elem)

        return self.markup

    def step_back(self, name_button):
        back = types.InlineKeyboardButton(text='<< назад', callback_data=name_button)

        return back

    def home(self):
        home = types.InlineKeyboardButton(text='<<< на главную', callback_data='home')

        return home

    def main_menu(self):
        keyboard = self.init_menu(self.list_main_menu)

        return keyboard

    def lvl_2(self, list_function):
        keyboard = self.init_menu(list_function)
        keyboard.add(self.home())

        return keyboard

    def lvl_3(self, list_function, name_button):
        keyboard = self.init_menu(list_function)
        keyboard.add(self.step_back(name_button))
        keyboard.add(self.home())

        return keyboard


class Exchange_with_ERP:
    """Позволяет получить некоторые данные из 1С"""

    def __init__(self, message, params):
        self.url = Data.way_erp
        self.login = Data.login_erp
        self.password = Data.pass_erp
        self.user_agent_val = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ' \
                              'Chrome/75.0.3770.142 Safari/537.36 '
        self.user_id = message.from_user.id
        self.user_name = message.from_user.first_name
        self.text_from_user = message.text
        self.params = params
        self.response = self.get_request()

    def get_request(self):
        request = requests.get(url=f'{self.url}',
                               headers={'User-Agent': self.user_agent_val},
                               auth=HTTPBasicAuth(self.login, self.password),
                               params=self.params)

        return request

    def answer_from_ERP(self):
        try:
            if self.response.status_code == 200:
                if Data.func_name1 in self.response.json():
                    return self.in_out()
                elif Data.func_name2 in self.response.json():
                    return self.get_count_days()
                elif Data.func_name3 in self.response.json():
                    return self
            elif self.response.status_code == 400:
                error_text = self.response.json().value
                return error_text
            elif self.response.status_code == 500:
                update_text = 'Не удалось выполнить запрос, возникла ошибка. Сервер 1С недоступен. ' \
                              'Производится обновление. Повторите попытку позже.'
                return update_text
        except Exception:
            exception_text = 'Не удалось выполнить запрос, возникла ошибка. Сервер 1С недоступен. ' \
                             'Возможно производится обновление. Повторите попытку позже.'
            return exception_text

    def get_count_days(self):
        """На вход принимает user_id, запрашивает данные из 1С, и возвращает кол-во накопленных дней отпуска.
        Если пользователь не уволен, функция вернёт число, во всех остальных случаях 1С вернёт ошибку"""

        json = self.response.json()
        count_day = int(json[Data.func_name2])

        logging_event('info',
                      f'Пользователь {self.user_name}({self.user_id}) получил ответ от ERP по кол-ву '
                      f'накопленных дней отпуска.')
        return count_day

    def verification(self):
        """Запрос принимает user_id и ИНН пользователя. В случае успеха, обновляет ID Telegram в 1С у пользователя с
        указанным ИНН. Либо возвращает str(ошибку)."""

        json = self.response.json()
        answer_erp = json[Data.func_name3]
        logging_event('info',
                      f'Пользователь {self.user_name}({self.user_id}) получил ответ от ERP по верификации.')
        return answer_erp

    def in_out(self):
        """Хз че делает"""

        json = self.response.json()
        data_list = []
        for value in json.values():
            for elem in value:
                in_out = elem['Вход']
                time = elem['Время']
                string_name = f'{time} {in_out}'
                data_list.append(string_name)

        return data_list
