import datetime
import urllib

from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler
import urllib.request

import Data


def read_file(list_name):
    opener = urllib.request.build_opener(SMBHandler)
    file_name = opener.open(Data.route)
    wb = load_workbook(file_name)  # Открываем нужную книгу
    sheet = wb[list_name]  # Получить лист по ключу

    # Получаем разницу между датами
    now_date = datetime.datetime.now()  # Получаем текущую дату
    some_date = sheet.cell(row=2, column=1).value  # Дата во 2й строке 1го столбца
    error = 'Ошибка в листе ' + list_name + ':\n'
    if some_date is not None:
        if isinstance(some_date, datetime.datetime):
            some_date = some_date
            some_date2 = sheet.cell(row=2, column=2).value  # Дата во 2й строке 2го столбца
            meaning = sheet.cell(row=2, column=3).value  # Текст во 2й строке 3го столбца
            meaning2 = sheet.cell(row=2, column=2).value  # Текст во 2й строке 2го столбца
            difference_date = some_date - now_date  # Получаем разницу между датами
            difference_date = difference_date.days  # Конвертируем дату в формат без учёта времени
            difference_date = int(difference_date) + 1
            read_type = 'date'
            read_data = {
                'Date 1': some_date,
                'Date 2': some_date2,
                'Text 3': meaning,
                'Text 2': meaning2,
                'Dif date': difference_date,
                'Type': read_type
            }
        else:
            error = error + 'Некорректные данные в поле "Дата". Необходимо указать в формате <дд.мм.гггг>' + '\n'
            read_type = 'incorrect'
            read_data = {
                'Error': error,
                'Type': read_type
            }
    elif some_date is None:
        error = error + 'Отсутствуют данные' + '\n'
        read_type = 'none'
        read_data = {
            'Date 1': error,
            'Dif date': error,
            'Error': error,
            'Type': read_type
        }

    return read_data
