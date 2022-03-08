import datetime
import os
import time
import urllib

from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler
import urllib.request

import Classes
import Count
import Data
import logging


def get_key(user_name):
    """Проверяет среди словаря есть ли в нём имя{user_name} и возвращает соответствующий id_telegram"""

    # key = 'Имя не найдено или человек не относится к дежурным. Error Other_function.get_key.'
    for keys, v in Data.user_data.items():
        for a in v:
            if a == user_name:
                key = keys
                return key


def get_data_user_SQL(user_name):
    """Принимает на вход Имя, проверяет есть ли это имя в списке дежурных, если есть, получает user_id, если он есть в
    БД, подтягивает Имя и user_name. Результат - <имя + @username>"""

    id_telegram = get_key(user_name)  # Присваиваем id из get_key
    print(id_telegram)
    if id_telegram is not None:  # Если id есть то
        if Classes.SQL().check_for_existence(id_telegram) is True:  # Если в SQL есть такой id
            end_text = Classes.SQL().get_user_info(id_telegram)  # Получаем склейку <имя + @username>
            print(end_text)
            return end_text


def logging_event(condition, text):
    if text is not None:
        logging.basicConfig(filename=Data.way_to_log_file, level=logging.INFO,
                            format="%(asctime)s - [%(levelname)s] - %(message)s")
        if condition == 'debug':
            logging.debug(text)
        elif condition == 'info':
            logging.info(text)
        elif condition == 'warning':
            logging.warning(text)
        elif condition == 'error':
            logging.error(text)
        elif condition == 'critical':
            logging.critical(text)


class File_processing:
    """Обработка файла"""

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.opener = urllib.request.build_opener(SMBHandler)
        self.file_name = self.opener.open(Data.route)
        self.wb = load_workbook(self.file_name)  # Открываем нужную книгу
        self.sheet = self.wb[sheet_name]  # Получить лист по имени
        self.now_date = datetime.datetime.now()  # Получаем текущую дату
        if (self.sheet.cell(row=1, column=1).value and
            self.sheet.cell(row=1, column=2).value and
            self.sheet.cell(row=1, column=3).value) is not None:  # Если заполнены 3 колонки

            self.count_meaning = 3  # Количество непустых колонок 3
        else:
            self.count_meaning = 2  # Количество непустых колонок 2

    def difference_date(self, date):
        f"""На вход принимает дату. Возвращает разницу в днях между сегодня и {date}. 
        Где "0" означает сегодня, "1" - "завтра" и тд"""

        difference = date - self.now_date  # Разница между 1‑й датой и сегодня
        difference = difference.days + 1  # Форматируем в кол-во дней +1

        return difference

    def read_file(self):
        """Возвращает списки значений в формате [ДД.ММ.ГГГГ, ДД.ММ.ГГГГ, событие]. Если дата в списке одна, то
        соответственно [ДД.ММ.ГГГГ, событие]."""

        self.clear_old_data()

        date_list = []  # Объявляем пустой список

        for i in range(1, self.sheet.max_row):  # Повторить для каждого значения в колонке А
            if i is not None:  # Если значение не пустое
                # print(f'Строка <{i}> с тексом <{self.sheet.cell(row=i, column=1).value}> '
                #       f'имеет тип <{type(self.sheet.cell(row=i, column=1).value)}>')
                if isinstance(self.sheet.cell(row=i, column=1).value, datetime.datetime):  # Если значение == дата
                    if self.count_meaning == 2:  # Если заполнены 2 колонки
                        date = self.sheet.cell(row=i, column=1).value  # Ячейка с датой
                        meaning = self.sheet.cell(row=i, column=2).value  # Ячейка с событием
                        line = [date, meaning]  # Лист из 2 значений
                        date_list.append([line])  # Добавить лист в общий список
                    elif self.count_meaning == 3:  # Если заполнены 3 колонки
                        date_one = self.sheet.cell(row=i, column=1).value  # Ячейка с первой датой
                        date_two = self.sheet.cell(row=i, column=2).value  # Ячейка со второй датой
                        meaning = self.sheet.cell(row=i, column=3).value  # Ячейка с событием
                        line = [date_one, date_two, meaning]  # Лист из 3 значений
                        date_list.append(line)  # Добавить лист в общий список

        date_list.sort()  # Сортируем лист по дате

        if len(date_list) == 0:  # Если список пуст
            print(f'Список <{self.sheet_name}> пуст')
            return None
        else:
            return date_list

    def clear_old_data(self):
        """Очистка неактуальных данных"""

        for i in range(1, self.sheet.max_row):  # Повторить для каждого значения в 1 колонке
            if i is not None:  # Если значение не пустое
                if isinstance(self.sheet.cell(row=i, column=1).value, datetime.datetime):  # Если значение == дата
                    if self.difference_date(self.sheet.cell(row=i, column=1).value) < 0:  # Если событие в прошлом
                        date_event = self.sheet.cell(row=i, column=1)  # Колонка с датой
                        event = self.sheet.cell(row=i, column=self.count_meaning).value  # Ячейка с событием
                        print(f'Удалена строка {date_event.row}\n'
                              f'Дата: {date_event.value}\n'
                              f'Текст: {event}\n\n')
                        self.sheet.delete_rows(i)  # Удаляем указанную в скобках строку
                        self.wb.save('test.xlsx')  # Сохранить книгу
                        file = open('test.xlsx', 'rb')
                        self.file_name = self.opener.open(Data.route, data=file)
                        self.file_name.close()
                        os.remove('test.xlsx')
                        time.sleep(1)
                        self.clear_old_data()
                    else:
                        continue

    def next_dej(self):
        """Если файл заполнен, возвращает строку 'В период с {first_date} по {second_date} будет дежурить {name}.'"""

        if self.sheet_name == 'Дежурный':
            self.clear_old_data()
            if self.read_file() is not None:
                name_from_SQL = Classes.SQL().get_user_info(get_key(self.read_file()[0][2]))
                text_message = f'В период с {self.read_file()[0][0].strftime("%d.%m.%Y")} ' \
                               f'по {self.read_file()[0][1].strftime("%d.%m.%Y")} ' \
                               f'будет дежурить {name_from_SQL}.'
                return text_message

    def sticker_next_dej(self):
        """Сравнивает имя из файла с БД, и если есть совпадение и в БД содержится стикер то вернёт его."""

        self.clear_old_data()
        if self.read_file() is not None:
            sticker = Classes.SQL().get_user_sticker(get_key(self.sheet.cell(row=2, column=self.count_meaning).value))
            if sticker is not None:
                return sticker
            else:
                return None

    def list_dej(self):
        """Возвращает список всех дежурных, что есть в списке"""

        list_dej = []
        for i in self.read_file():
            if self.sheet_name == 'Дежурный':
                if i is not None:
                    date_one = i[0].strftime("%d.%m.%Y")
                    date_two = i[1].strftime("%d.%m.%Y")
                    name_from_SQL = Classes.SQL().get_user_info(get_key(i[2]))
                    text_message = f'В период с {date_one} по {date_two} будет дежурить {name_from_SQL}.'
                    list_dej.append(text_message)

        return list_dej

    def next_invent(self):
        """Если файл заполнен, возвращает строку 'До предстоящей инвентаризации осталось {N} дней.
        Судя по графику, выходит {name}'"""

        if self.sheet_name == 'Инвентаризация':
            if self.read_file() is not None:
                name_from_SQL = Classes.SQL().get_user_info(get_key(self.sheet.cell(row=2, column=2).value))
                dead_line_date = self.difference_date(self.sheet.cell(row=2, column=1).value)
                dead_line_text = Count.Counter().days_before_inventory(dead_line_date)
                text_who = f'Судя по графику, выходит {name_from_SQL}.'  # Имя следующего дежурного
                text_message = f'{dead_line_text} {text_who}'
                return text_message
            elif self.read_file() is None:
                text_message = 'Нет данных об этом. Необходимо заполнить файл!'
                return text_message

    def check_event_today(self):
        """Проверяет есть ли сегодня событие. Результат - уведомление соответствующей группе пользователей."""

        if self.read_file() is not None:
            for i in range(1, self.sheet.max_row):
                if self.sheet_name in Data.sheets_file:
                    if isinstance(self.sheet.cell(row=i, column=1).value, datetime.datetime):  # Если значение == дата
                        if self.difference_date(self.sheet.cell(row=i, column=1).value) == 0:
                            if self.sheet_name == 'Уведомления для всех':
                                text_message = f'• Уведомление для зарегистрированных пользователей •\n\n' \
                                               f'{self.sheet.cell(row=i, column=2).value}'
                                print(text_message)
                                Classes.Notification().all_registered(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для подписчиков':
                                text_message = f'• Уведомление для подписчиков •\n\n' \
                                               f'{self.sheet.cell(row=i, column=2).value}'
                                print(text_message)
                                Classes.Notification().notification_for(text_message, 'notification', 'yes')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для админов':
                                text_message = f'• Уведомление для администраторов •\n\n' \
                                               f'{self.sheet.cell(row=i, column=2).value}'
                                print(text_message)
                                Classes.Notification().notification_for(text_message, 'status', 'admin')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                        # elif self.difference_date(self.sheet.cell(row=i, column=1).value) == 1:
                        #     if self.sheet_name == 'Дежурный':
                        #         text_message = self.next_dej()
                        #         name_dej = self.sheet.cell(row=i, column=3).value
                        #         print(text_message)
                        #         Classes.Notification().notification_for(text_message, 'notification', 'yes')
                        #         Classes.Notification().send_sticker_for(name_dej, 'notification', 'yes')
                        #         # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                        #     elif self.sheet_name == 'Инвентаризация':
                        #         self.next_invent()
                        elif self.difference_date(
                                self.sheet.cell(row=i, column=1).value) >= 1:  # Если дата не наступила
                            if self.count_meaning == 2:
                                print(f'{File_processing.check_event_today.__qualname__}\n'
                                      f'Событие не наступило\nЛист:{self.sheet_name}\n'
                                      f'Текст уведомления:{self.sheet.cell(row=i, column=2).value}\n'
                                      f'Дата:{self.sheet.cell(row=i, column=1).value.strftime("%d.%m.%Y")}\n\n')
                            elif self.count_meaning == 3:
                                print(f'{File_processing.check_event_today.__qualname__}\n'
                                      f'Событие не наступило\nЛист:{self.sheet_name}\n'
                                      f'Текст уведомления:{self.sheet.cell(row=i, column=3).value}\n'
                                      f'Дата:{self.sheet.cell(row=i, column=1).value.strftime("%d.%m.%Y")}\n\n')

                        time.sleep(5)

    def check_dej_today(self):
        """Проверяет есть ли завтра дежурный. Результат - уведомление соответствующей группе пользователей."""

        if self.read_file() is not None:
            for i in range(1, self.sheet.max_row):
                if self.sheet_name in Data.sheets_file:
                    if isinstance(self.sheet.cell(row=i, column=1).value, datetime.datetime):  # Если значение == дата
                        if self.difference_date(self.sheet.cell(row=i, column=1).value) == 1:
                            if self.sheet_name == 'Дежурный':
                                text_message = self.next_dej()
                                name_dej = self.sheet.cell(row=i, column=3).value
                                print(text_message)
                                Classes.Notification().notification_for(text_message, 'notification', 'yes')
                                Classes.Notification().send_sticker_for(name_dej, 'notification', 'yes')
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Инвентаризация':
                                self.next_invent()
                        elif self.difference_date(self.sheet.cell(row=i, column=1).value) >= 1:
                            '''Если дата не наступила'''
                            if self.count_meaning == 2:
                                print(f'{File_processing.check_event_today.__qualname__}\n'
                                      f'Событие не наступило\nЛист:{self.sheet_name}\n'
                                      f'Текст уведомления:{self.sheet.cell(row=i, column=2).value}\n'
                                      f'Дата:{self.sheet.cell(row=i, column=1).value.strftime("%d.%m.%Y")}\n\n')
                            elif self.count_meaning == 3:
                                print(f'{File_processing.check_event_today.__qualname__}\n'
                                      f'Событие не наступило\nЛист:{self.sheet_name}\n'
                                      f'Текст уведомления:{self.sheet.cell(row=i, column=3).value}\n'
                                      f'Дата:{self.sheet.cell(row=i, column=1).value.strftime("%d.%m.%Y")}\n\n')

    def create_event(self, date, text_event):
        """Принимает дату и текст уведомления и записывает в файл. Возвращает текст с описанием созданного
        уведомления """

        # i = 0
        try:
            for i in range(1, self.sheet.max_row):  # Повторить для каждого значения в колонке А
                if self.sheet.cell(row=i, column=1).value is None:
                    print(f'Строка {self.sheet.cell(row=i, column=1).row} пустая')
                    empty_string = self.sheet.cell(row=i, column=1).row
                    date_obj = datetime.datetime.strptime(date, '%d.%m.%Y')
                    self.sheet.cell(row=empty_string, column=1).value = date_obj  # Дата
                    self.sheet.cell(row=empty_string, column=2).value = text_event  # Текст события
                    self.wb.save('test.xlsx')  # Сохранить книгу
                    file = open('test.xlsx', 'rb')
                    self.file_name = self.opener.open(Data.route, data=file)
                    self.file_name.close()
                    os.remove('test.xlsx')
                    text_message = f'• Запись добавлена в лист: <{self.sheet_name}>\n' \
                                   f'• Дата уведомления: <{date}>\n' \
                                   f'• Текст: <{text_event}>\n'

                    return text_message
        except urllib as e:
            print(f'Ошибка при работе с файлом:\n{e}')
        except Exception as e:
            # time.sleep(3)
            print(str(e))
            # Other_function.logging_event('error', str(e))
