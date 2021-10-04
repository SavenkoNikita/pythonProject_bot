import datetime
import os
import urllib.request

from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler

import Data
import Read_file


def clear(list_name):
    # comment = ''
    opener = urllib.request.build_opener(SMBHandler)
    file_name = opener.open(Data.route)
    wb = load_workbook(file_name)  # Открываем нужную книгу
    sheet = wb[list_name]  # Получить лист по имени

    some_date = sheet.cell(row=2, column=1).value  # Извлекаем дату из файла

    if some_date is not None:
        if isinstance(some_date, datetime.datetime):
            some_date = some_date
            now_date = datetime.datetime.now()  # Получаем текущую дату
            difference_date = some_date - now_date  # Получаем разницу между датами
            difference_date = difference_date.days  # Конвертируем дату в формат без учёта времени
            difference_date = int(difference_date) + 1
            if difference_date < 0:  # Если дата в прошлом
                sheet.delete_rows(2)  # Удаляем указанную в скобках строку
                wb.save('test.xlsx')  # Сохранить книгу
                file = open('test.xlsx', 'rb')
                file_name = opener.open(Data.route, data=file)
                file_name.close()
                os.remove('test.xlsx')
                del_data = '• Из списка ' + str(list_name) + ' удалены данные'
                # comment = comment + del_data
                print(del_data)
            else:
                actual_data = '• В списке <' + str(list_name) + '> все данные актуальны'
                # comment = comment + actual_data
                print(actual_data)
    elif some_date is None:
        none_data = '• В списке <' + str(list_name) + '> нет данных'
        # comment = comment + none_data
        print(none_data)


# Проверка данных на релевантность
def check_relevance(list_name):
    difference_date = Read_file.read_file(list_name)['Dif date']  # Извлекаем ближайшую дату из list_name
    while difference_date < 0:  # Повторить если дата в прошлом
        # difference_date = Read_file.read_file(list_name)['Dif date']
        clear(list_name)
    else:
        print('Данные релевантны')
