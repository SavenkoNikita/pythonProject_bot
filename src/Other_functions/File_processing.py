import datetime
import os
import time
import traceback
import urllib

from openpyxl import load_workbook
from smb.SMBHandler import SMBHandler

import Data
from src.Other_functions.Functions import SQL, get_key, days_before_inventory, logging_file_processing
from src.Other_functions.Working_with_notifications import Notification
from urllib.request import urlopen


class Working_with_a_file:
    """Обработка файла"""

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.opener = urllib.request.build_opener(SMBHandler)
        self.file_name = self.opener.open(Data.route)
        self.wb = load_workbook(self.file_name)  # Открываем нужную книгу
        self.sheet = self.wb[sheet_name]  # Получить лист по имени
        self.first_column = self.sheet['A']
        self.now_date = datetime.datetime.now()  # Получаем текущую дату
        self.title_first_column = self.sheet.cell(row=1, column=1).value  # Заголовок первой колонки
        self.title_second_column = self.sheet.cell(row=1, column=2).value  # Заголовок второй колонки
        self.title_third_column = self.sheet.cell(row=1, column=3).value  # Заголовок третьей колонки
        if (self.title_first_column and self.title_second_column and self.title_third_column) is not None:
            """Если заполнены 3 колонки, {self.count_meaning} возвращает '3', а если 2, то '2'."""
            self.count_meaning = 3  # Количество непустых колонок 3
        else:
            self.count_meaning = 2  # Количество непустых колонок 2

    def difference_date(self, date):
        f"""На вход принимает дату. Возвращает int(разницу) в днях между сегодня и {date}. 
        Где "0" означает сегодня, "1" - "завтра" и тд"""

        difference = date - self.now_date  # Разница между 1‑й датой и сегодня
        difference = difference.days + 1  # Форматируем в кол-во дней +1

        if difference == -1:
            print('Событие было вчера')
        elif difference == 0:
            print('Событие сегодня')
        elif difference == 1:
            print('Событие завтра')

        text_to_log = f'difference_date(): ' \
                      f'Передана дата {date}. ' \
                      f'Текущая дата {self.now_date}. ' \
                      f'Разница между датами = {difference}.'
        logging_file_processing('info', text_to_log)
        # print(text_to_log)

        return difference

    def read_file(self):
        """Возвращает списки значений в формате [ДД.ММ.ГГГГ(datetime), ДД.ММ.ГГГГ(datetime), событие].
        Если дата в списке одна, то соответственно [ДД.ММ.ГГГГ(datetime), событие]."""

        self.clear_old_data()

        date_list = []  # Объявляем пустой список

        try:
            for i in range(2, 100):  # Повторить для каждого значения в колонке А
                if self.sheet.cell(row=i, column=1).value is not None:  # Если ячейка не пустая
                    if self.count_meaning == 2:  # Если заполнены 2 колонки
                        date = self.sheet.cell(row=i, column=1).value  # Ячейка с датой
                        if isinstance(date, datetime.datetime):  # Если значение == дата
                            date = date
                        else:
                            date = datetime.datetime.strptime(date, '%d.%m.%Y')
                        meaning = self.sheet.cell(row=i, column=2).value  # Ячейка с событием
                        line = [date, meaning]  # Лист из 2 значений
                        date_list.append(line)  # Добавить лист в общий список
                    elif self.count_meaning == 3:  # Если заполнены 3 колонки
                        date_one = self.sheet.cell(row=i, column=1).value  # Ячейка с первой датой
                        if isinstance(date_one, datetime.datetime):  # Если значение == дата
                            date_one = date_one
                        else:
                            date_one = datetime.datetime.strptime(date_one, '%d.%m.%Y')

                        date_two = self.sheet.cell(row=i, column=2).value  # Ячейка со второй датой
                        if isinstance(date_two, datetime.datetime):  # Если значение == дата
                            date_two = date_two
                        else:
                            date_two = datetime.datetime.strptime(date_two, '%d.%m.%Y')

                        meaning = self.sheet.cell(row=i, column=3).value  # Ячейка с событием
                        line = [date_one, date_two, meaning]  # Лист из 3 значений
                        date_list.append(line)  # Добавить лист в общий список

            text_log = f'read_file(): Данные сформированы и переданы из листа {self.sheet_name}'
            logging_file_processing('info', text_log)
        except Exception.__traceback__:
            error = traceback.format_exc()
            logging_file_processing('debug', error)
            print(error)
        finally:
            date_list.sort()  # Сортируем лист по дате
            # print(date_list)

            if len(date_list) == 0:  # Если список пуст
                no_data_available = f'read_file(): Список <{self.sheet_name}> пуст'
                logging_file_processing('info', no_data_available)
                print(f'{no_data_available}\n')
                return None
            else:
                read_file = f'read_file(): Прочитаны и возвращены данные листа <{self.sheet_name}>'
                logging_file_processing('info', read_file)
                # print(date_list)
                return date_list

    def clear_old_data(self):
        """Очистка неактуальных данных"""

        for i in range(2, 10):  # Повторить для каждого значения в 1 колонке
            date = self.sheet.cell(row=i, column=1).value
            # print(date)
            # print(type(date))
            if date is not None:  # Если значение не пустое
                if isinstance(date, datetime.datetime):  # Если значение == дата
                    # print('date')
                    date = date
                else:
                    date = datetime.datetime.strptime(date, '%d.%m.%Y')
                    # print(f'{date} not date')

                if self.difference_date(date) < 0:  # Если событие в прошлом
                    date_event = self.sheet.cell(row=i, column=1)  # Колонка с датой
                    event = self.sheet.cell(row=i, column=self.count_meaning).value  # Ячейка с событием
                    old_data = f'Со страницы <{self.sheet_name}>\n'\
                               f'Удалена строка {date_event.row}\n' \
                               f'Дата: {date_event.value}\n' \
                               f'Текст: {event}\n\n'
                    logging_file_processing('info', old_data)
                    print(old_data)
                    self.sheet.delete_rows(i)  # Удаляем указанную в скобках строку
                    self.wb.save('test.xlsx')  # Сохранить книгу
                    file = open('test.xlsx', 'rb')
                    self.file_name = self.opener.open(Data.route, data=file)  # noqa
                    self.file_name.close()
                    os.remove('test.xlsx')
                    time.sleep(1)
                    self.clear_old_data()
                    # else:
                    #     print(f'{self.difference_date(self.sheet.cell(row=i, column=1).value)}')
                # else:
                #     print(f'{type(self.sheet.cell(row=i, column=1).value)} не дата')
            # else:
            #     print(f'{self.sheet.cell(row=i, column=1).value} is None')

    def next_dej(self):
        """Если файл заполнен, возвращает строку 'В период с {first_date} по {second_date} будет дежурить {name}.'"""

        data_list = self.read_file()

        if self.sheet_name == 'Дежурный':
            self.clear_old_data()
            if data_list is not None:
                first_date = data_list[0][0].strftime('%d.%m.%Y')  # Дата str(1)
                second_date = data_list[0][1].strftime('%d.%m.%Y')  # Дата str(2)
                name_from_SQL = SQL().get_user_info(get_key(self.read_file()[0][2]))  # Имя дежурного

                text_message = f'В период с {first_date} по {second_date} будет дежурить {name_from_SQL}.'
                logging_file_processing('info', text_message)

                return text_message
            else:
                text_message = 'Нет данных о предстоящих дежурствах :('

                return text_message

    def sticker_next_dej(self):
        """Сравнивает имя из файла с БД, и если есть совпадение и в БД содержится стикер то вернёт его."""

        # self.clear_old_data()
        if self.read_file() is not None:
            sticker = SQL().get_a_user_sticker_from_the_database(
                get_key(self.sheet.cell(row=2, column=self.count_meaning).value))
            if sticker is not None:
                text_log = f'sticker_next_dej(): Возвращён стикер пользователя <{sticker}>.'
                logging_file_processing('info', text_log)
                return sticker
            else:
                text_log = f'sticker_next_dej(): Пользователь {self.sheet.cell(row=2, column=self.count_meaning)} ' \
                           f'не имеет стикера в БД. Результат функции - None'
                logging_file_processing('info', text_log)
                return None

    def list_dej(self):
        """Возвращает список всех дежурных, что есть в списке"""

        list_dej = []
        for i in self.read_file():
            if self.sheet_name == 'Дежурный':
                if i is not None:
                    date_one = i[0].strftime("%d.%m.%Y")
                    date_two = i[1].strftime("%d.%m.%Y")
                    name_from_SQL = SQL().get_user_info(get_key(i[2]))
                    text_message = f'В период с {date_one} по {date_two} будет дежурить {name_from_SQL}.'
                    list_dej.append(text_message)
                else:
                    text_message = 'Нет данных о предстоящих дежурствах :('

                    return text_message

        text_log = f'list_dej(): {list_dej}'
        logging_file_processing('info', text_log)
        return list_dej

    def next_invent(self):
        """Если файл заполнен, возвращает строку 'До предстоящей инвентаризации осталось {N} дней.
        Судя по графику, выходит {name}'."""

        if self.read_file() is not None:
            date_event = self.read_file()[0][0]
            name_in_list = self.read_file()[0][1]
            name_from_SQL = SQL().get_user_info(get_key(name_in_list))
            if name_from_SQL is None:
                name_from_SQL = name_in_list
            dead_line_date = self.difference_date(date_event)
            dead_line_text = days_before_inventory(dead_line_date)
            text_who = f'Судя по графику, выходит {name_from_SQL}.'  # Имя следующего дежурного
            text_message = f'{dead_line_text} {text_who}'
            text_log = f'next_invent(): {text_message}'
            logging_file_processing('info', text_log)
            return text_message
        elif self.read_file() is None:
            text_message = f'Лист {self.sheet_name} пуст. Необходимо заполнить файл!'
            logging_file_processing('info', text_message)
            return text_message

    def check_event_today(self):
        """Проверяет есть ли сегодня событие. Результат - уведомление соответствующей группе пользователей."""

        data_list = self.read_file()

        if data_list is not None:  # Если лист не пуст
            # print(f'data_list not None')
            if self.sheet_name in Data.sheets_file:  # Если название листа есть в списке известных
                for i in data_list:
                    date = i[0]
                    # print(date)
                    meaning = i[1]
                    # print(meaning)
                    if len(i) == 2:
                        if self.difference_date(date) == 0:
                            if self.sheet_name == 'Уведомления для всех':
                                text_message = f'• Уведомление для зарегистрированных пользователей •\n\n' \
                                               f'{meaning}'
                                text_log = f'check_event_today(): {text_message}'
                                logging_file_processing('info', text_log)
                                print(text_message)
                                Notification().send_a_notification_to_all_users(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для подписчиков':
                                text_message = f'• Уведомление для подписчиков •\n\n' \
                                               f'{meaning}'
                                text_log = f'check_event_today(): {text_message}'
                                logging_file_processing('info', text_log)
                                print(text_message)
                                Notification().send_notification_to_subscribers(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для админов':
                                text_message = f'• Уведомление для администраторов •\n\n' \
                                               f'{meaning}'
                                text_log = f'check_event_today(): {text_message}'
                                logging_file_processing('info', text_log)
                                print(text_message)
                                Notification().send_notification_to_administrators(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                            elif self.sheet_name == 'Уведомления для барахолки':
                                text_message = f'• Уведомление для подписчиков барахолки •\n\n' \
                                               f'{meaning}'
                                text_log = f'check_event_today(): {text_message}'
                                logging_file_processing('info', text_log)
                                print(text_message)
                                Notification().notification_for_sub_baraholka(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                        elif 0 <= self.difference_date(date) <= 2:
                            if self.sheet_name == 'Инвентаризация':
                                text_message = f'• Уведомление для администраторов •\n\n' \
                                               f'{self.next_invent()}'
                                text_log = f'check_event_today(): {text_message}'
                                logging_file_processing('info', text_log)
                                print(text_message)
                                Notification().send_notification_to_administrators(text_message)
                                # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)

                        time.sleep(5)

    def check_dej_tomorrow(self):
        """Проверяет есть ли завтра дежурный. Результат - уведомление соответствующей группе пользователей."""

        data_list = self.read_file()

        if data_list is not None:
            for i in data_list:
                date = i[0]
                if self.difference_date(date) == 1:  # Если завтра
                    first_date = i[0].strftime('%d.%m.%Y')  # Дата str(1)
                    second_date = i[1].strftime('%d.%m.%Y')  # Дата str(2)
                    name = i[2]
                    name_from_SQL = SQL().get_user_info(get_key(i[2]))  # Имя дежурного

                    text_message = f'В период с {first_date} по {second_date} будет дежурить {name_from_SQL}.'

                    text_log = f'check_dej_tomorrow(): {text_message}'
                    logging_file_processing('info', text_log)
                    print(f'<{datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}>\n{text_message}')

                    sticker_dej = SQL().get_a_user_sticker_from_the_database(get_key(name))

                    Notification().send_notification_to_subscribers(text_message)
                    Notification().send_sticker_to_subscribers(sticker_dej)
                    # Data.bot.send_message(Data.list_admins.get('Никита'), text_message)
                    # Data.bot.send_sticker(Data.list_admins.get('Никита'), sticker_dej)
                elif self.difference_date(date) < 0:  # Если событие в прошлом
                    text_log = f'check_dej_tomorrow(): Событие в прошлом.'
                    logging_file_processing('info', text_log)
                    self.clear_old_data()
                # else:
                #     print(f'Дата {date} не наступила')

    def create_event(self, date, text_event):
        """Принимает дату и текст уведомления и записывает в файл. Возвращает текст с описанием созданного
        уведомления """

        try:
            for i in range(1, 1000):  # Повторить для каждого значения в колонке А
                if self.sheet.cell(row=i, column=1).value is None:
                    print(f'Строка {self.sheet.cell(row=i, column=1).row} пустая')
                    empty_string = self.sheet.cell(row=i, column=1).row
                    date_obj = datetime.datetime.strptime(date, '%d.%m.%Y')
                    date_obj = date_obj.date().strftime('%d.%m.%Y')
                    self.sheet.cell(row=empty_string, column=1).value = date_obj  # Дата
                    self.sheet.cell(row=empty_string, column=2).value = text_event  # Текст события
                    self.wb.save('test.xlsx')  # Сохранить книгу
                    file = open('test.xlsx', 'rb')
                    self.file_name = self.opener.open(Data.route, data=file)  # noqa
                    self.file_name.close()
                    os.remove('test.xlsx')
                    text_message = f'• Запись добавлена в лист: <{self.sheet_name}>\n' \
                                   f'• Дата уведомления: <{date_obj}>\n' \
                                   f'• Текст: <{text_event}>\n'

                    text_log = f'create_event(): {text_message}'
                    logging_file_processing('info', text_log)

                    return text_message
        except urllib as e:
            text_error = f'create_event(): Ошибка при работе с файлом:\n{e}'
            logging_file_processing('debug', text_error)
            print(text_error)
        except Exception as e:
            time.sleep(3)
            text_error = f'create_event(): Ошибка при работе с файлом:\n{e}'
            logging_file_processing('debug', text_error)
            print(str(e))
